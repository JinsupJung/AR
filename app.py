# app.py

from flask import Flask, request, render_template, redirect, url_for, flash, jsonify, send_from_directory
from flask_wtf import FlaskForm
from wtforms import SelectField, StringField, SubmitField, DecimalField, DateField, FileField
from wtforms.validators import DataRequired, NumberRange
from werkzeug.utils import secure_filename
import pandas as pd
import mysql.connector
import logging 
import os
from datetime import datetime
from flask_wtf.csrf import CSRFProtect
import json  # JSON 처리를 위해 추가
from decimal import Decimal  # Decimal 처리를 위해 추가
import re
import jaydebeapi
import traceback
from dotenv import load_dotenv
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from zipfile import ZipFile
import tempfile
import subprocess
import uuid
import threading
# PDF 병합용
from PyPDF2 import PdfMerger
from flask import send_file
# PDF 출력용 0107 추가
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# .env 파일 로드 (보안을 위해 환경 변수 사용 권장)
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your_default_secret_key')  # 보안을 위해 환경 변수 사용 권장

# CSRF 보호 설정
csrf = CSRFProtect(app)

# 파일 업로드 설정
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 최대 업로드 크기 설정 (16MB)

# 로그 설정
today = datetime.now().strftime("%Y%m%d")
log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
os.makedirs(log_dir, exist_ok=True)
log_filename = os.path.join(log_dir, f'app_{today}.log')

logging.basicConfig(
    level=logging.DEBUG,  # 로깅 레벨을 DEBUG로 설정
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# 전역 딕셔너리: 백그라운드 작업 결과 저장 (task_id -> 결과 파일 경로 리스트)
background_tasks = {}

# MySQL 연결 설정 함수
def get_db_connection():
    try:
        db = mysql.connector.connect(
            host=os.getenv('DB_HOST', '175.196.7.45'),
            user=os.getenv('DB_USER', 'nolboo'),
            password=os.getenv('DB_PASSWORD', '2024!puser'),
            database=os.getenv('DB_NAME', 'nolboo'),
            charset='utf8mb4'
        )
        logging.info("MySQL 데이터베이스에 성공적으로 연결되었습니다.")
        return db
    except mysql.connector.Error as err:
        logging.error(f"MySQL 연결 오류: {err}")
        return None

# 프로그램 설정
TEMPLATE_FILE = 'detail_form.xlsx'  # 엑셀 템플릿 파일명
OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')  # 출력 엑셀 파일 저장 디렉토리

# 테이블 이름
AR_ORDER_DETAILS_ITEM_TABLE = 'AROrderDetailsItem'
CM_CHAIN_TABLE = 'cm_chain'

# 엑셀 헤더와 MySQL 컬럼 매핑
COLUMN_MAPPING = {
    '배송일자': 'order_date',
    '거래처': 'rep_code',
    '거래처명': 'rep_name',
    '매출처': 'client_code',
    '매출처명': 'client_name',
    '제품코드': 'item_code',
    '제품명': 'item_name',
    '온도': 'cond',
    '출고단위': 'unit',
    '수량': 'qty',
    '계근 중량': 'cal_qty',
    '단가': 'unit_price',
    '공급가': 'order_amount',
    '부가세': 'vat',
    '합계': 'total_amount',
    '세금': 'tax'
}

# 공급자 정보 (고정)
SUPPLIER_INFO = {
    '등록번호': '112-81-22058',
    '상호 (법인명)': '(주) 놀부',
    '성명': '김용위',
    '주소': '서울특별시 강남구 영동대로 701, W타워 14~15층'
}

# 헬퍼 함수: 파일 확장자 확인
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 폼 정의
class UploadForm(FlaskForm):
    file = FileField('Upload Excel File', validators=[DataRequired()])
    submit = SubmitField('Upload')

class AddOrderForm(FlaskForm):
    client_code = SelectField('매출처', choices=[], validators=[DataRequired()])
    representative_code = StringField('대표코드', render_kw={'readonly': True})
    order_date = DateField('발주일', format='%Y-%m-%d', validators=[DataRequired()])
    amount = DecimalField('금액', validators=[DataRequired(), NumberRange(min=0)])
    submit = SubmitField('발주저장')

class BulkUploadOrdersForm(FlaskForm):
    file = FileField('발주 내역 엑셀 업로드', validators=[DataRequired()])
    submit = SubmitField('업로드')

# 새로운 폼 클래스 추가 (DailyTransactionsForm)
class DailyTransactionsForm(FlaskForm):
    current_year = datetime.now().year
    years = [(str(year), str(year)) for year in range(current_year - 5, current_year + 1)]
    months = [(str(month).zfill(2), str(month)) for month in range(1, 13)]
    
    year = SelectField('연도', choices=years, validators=[DataRequired()])
    month = SelectField('월', choices=months, validators=[DataRequired()])
    submit = SubmitField('조회')    

# 새로운 폼 클래스 정의
class DownloadOrdersForm(FlaskForm):
    order_date = DateField('처리할 배송일자', format='%Y-%m-%d', validators=[DataRequired()])
    submit = SubmitField('엑셀 다운로드')
# 신규 폼 클래스: DownloadClientOrdersForm

class DownloadClientOrdersForm(FlaskForm):
    # 필수 검증기 제거하여 전체매출처(빈 값)도 허용
    client_code = SelectField('매출처', choices=[('', '전체매출처')])
    from_date = DateField('시작일', format='%Y-%m-%d', validators=[DataRequired()])
    to_date = DateField('종료일', format='%Y-%m-%d', validators=[DataRequired()])
    submit = SubmitField('거래명세표 생성')

# ------------------------
# Informix 연결 정보 설정
# ------------------------
informix_jdbc_driver_class = 'com.informix.jdbc.IfxDriver'
informix_hostname = os.getenv('INFORMIX_HOST', '175.196.7.17')
informix_port = os.getenv('INFORMIX_PORT', '1526')
informix_database = os.getenv('INFORMIX_DATABASE', 'nolbooco')
informix_server = os.getenv('INFORMIX_SERVER', 'nbmain')
informix_username = os.getenv('INFORMIX_USERNAME', 'informix')
informix_password = os.getenv('INFORMIX_PASSWORD', 'eusr2206')  # 보안을 위해 환경 변수 사용 권장
jdbc_driver_path = os.getenv('JDBC_DRIVER_PATH', '/opt/IBM/Informix_JDBC_Driver/lib/ifxjdbc.jar')

informix_jdbc_url = (
    f"jdbc:informix-sqli://{informix_hostname}:{informix_port}/{informix_database}:"
    f"INFORMIXSERVER={informix_server};DBLOCALE=en_US.819;CLIENT_LOCALE=en_us.utf8;"
)

# ------------------------
# 3. 데이터 변환 함수 정의 (ETL용)
# ------------------------
def convert_to_utf8(value):
    if isinstance(value, str):
        try:
            temp_byte = value.encode('ISO-8859-1')  # 원본 인코딩에 맞게 수정 필요
            utf8_value = temp_byte.decode('euc-kr')  # Informix 데이터가 EUC-KR 인코딩이라면
            return utf8_value
        except Exception as e:
            logging.error(f"Failed to decode value '{value}': {e}")
            return value  # 디코딩 실패 시 원본 값 반환
    return value

def check_special_characters(df, columns):
    pattern = re.compile(r'[^\x00-\x7F]+')  # ASCII 외 문자 패턴
    for col in columns:
        if col in df.columns:
            problematic_rows = df[df[col].apply(lambda x: bool(pattern.search(x)) if isinstance(x, str) else False)]
            if not problematic_rows.empty:
                logging.warning(f"컬럼 '{col}'에 특수 문자가 포함된 데이터가 존재합니다.")
                logging.info(problematic_rows[[col]].to_string(index=False))
        else:
            logging.warning(f"'{col}' 컬럼이 데이터프레임에 존재하지 않습니다.")

def extract_data(cursor, query):
    cursor.execute(query)
    data = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    return pd.DataFrame(data, columns=columns)

def save_to_excel(df, path):
    df.to_excel(path, index=False)
    logging.info(f"데이터 엑셀로 저장 완료: {path}")

def log_query_string(query):
    non_ascii = [char for char in query if ord(char) > 127]
    if non_ascii:
        logging.warning("쿼리 문자열에 비ASCII 문자가 포함되어 있습니다:")
        logging.warning(''.join(non_ascii))
    else:
        logging.info("쿼리 문자열에 비ASCII 문자가 없습니다.")

# ------------------------
# 6. ETL 프로세스 함수 정의
# ------------------------
def fetch_client_data(engine, from_date, to_date, client_code):
    try:
        query = f"""
            SELECT 
                a.order_date,
                a.rep_code,
                a.rep_name,
                a.client_code,
                a.client_name,
                a.item_code,
                a.item_name,
                a.cond,
                a.unit,
                a.qty,
                a.cal_qty,
                a.unit_price,
                a.order_amount,
                a.vat,
                a.total_amount,
                a.tax,
                c.full_name,       -- 매출처명 (cm_chain의 full_name)
                c.reg_no,          -- 매출처 등록번호
                c.president,       -- 매출처 대표자
                c.address1         -- 매출처 주소
            FROM {AR_ORDER_DETAILS_ITEM_TABLE} a
            INNER JOIN {CM_CHAIN_TABLE} c 
              ON a.client_code COLLATE utf8mb4_unicode_ci = c.chain_no COLLATE utf8mb4_unicode_ci
            WHERE a.order_date BETWEEN '{from_date}' AND '{to_date}'
        """
        if client_code and client_code.strip() != "":
            query += f" AND a.client_code = '{client_code.strip()}'"
        query += " ORDER BY a.order_date, a.client_code"
        logging.debug(f"fetch_client_data 쿼리: {query}")
        df = pd.read_sql(query, engine)
        logging.info(f"MySQL에서 데이터를 성공적으로 조회했습니다. 총 {len(df)}개의 레코드.")
        return df
    except Exception as e:
        logging.error(f"매출처 데이터 조회 중 오류 발생: {e}", exc_info=True)
        return None

# def export_client_orders_to_files(from_date, to_date, client_code):
#     logging.debug("export_client_orders_to_files 함수 시작 (일자별 출력)")
#     logging.info(f"매개변수: from_date={from_date}, to_date={to_date}, client_code='{client_code}'")
    
#     engine = get_sqlalchemy_engine()
#     if engine is None:
#         logging.error("SQLAlchemy 엔진 생성 실패")
#         raise ConnectionError("SQLAlchemy 엔진 생성 실패")
    
#     df = fetch_client_data(engine, from_date, to_date, client_code)
#     logging.info(f"fetch_client_data 완료: {len(df)} 행 조회됨")
#     if df.empty:
#         logging.error("조회된 데이터가 없습니다.")
#         raise ValueError("조회된 데이터가 없습니다.")
    
#     try:
#         df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce').dt.strftime('%Y-%m-%d')
#         logging.debug("데이터 전처리 완료")
#     except Exception as e:
#         logging.error(f"데이터 전처리 오류: {e}", exc_info=True)
#         raise e

#     file_paths = []
#     # 항상 'order_date'와 'client_code'로 그룹화합니다.
#     grouped = df.groupby(['order_date', 'client_code'])
    
#     for group_keys, group in grouped:
#         day_str, client_grp = group_keys
#         logging.info(f"처리 그룹: order_date={day_str}, client_code={client_grp}")
#         try:
#             wb, ws = load_excel_template()
#             client_name = group['client_name'].iloc[0] if 'client_name' in group.columns else "unknown"
#             excel_path = generate_excel_file(wb, ws, client_name, day_str, group)
#             file_paths.append(excel_path)
#             logging.info(f"{day_str} 엑셀 파일 생성 완료: {excel_path}")
#             try:
#                 pdf_path = convert_excel_to_pdf(excel_path, OUTPUT_FOLDER)
#                 file_paths.append(pdf_path)
#                 logging.info(f"{day_str} PDF 파일 생성 완료: {pdf_path}")
#             except Exception as e:
#                 logging.error(f"PDF 변환 오류 ({day_str}): {e}", exc_info=True)
#         except Exception as e:
#             logging.error(f"엑셀 파일 생성 오류 (order_date={day_str}, client_code={client_grp}): {e}", exc_info=True)
    
#     logging.debug("export_client_orders_to_files 함수 종료 (일자별 출력)")
#     return file_paths

# def generate_excel_file(wb, ws, client_name, order_date, group):
#     try:
#         client_info = {
#             'full_name': group['full_name'].iloc[0],
#             'reg_no': group['reg_no'].iloc[0],
#             'president': group['president'].iloc[0],
#             'address1': group['address1'].iloc[0]
#         }
#         insert_data_to_excel(wb, ws, SUPPLIER_INFO, client_info, order_date, group.to_dict('records'))
#         sanitized_client_name = re.sub(r'[\\/*?:"<>|]', "_", client_name)  # 파일명에 사용할 수 없는 문자를 _로 대체
#         sanitized_order_date = order_date.replace('-', '')
#         if sanitized_client_name == "":
#             sanitized_client_name = "unknown"
#         output_filename = f"거래명세표_{sanitized_client_name}_{sanitized_order_date}.xlsx"
#         output_path = os.path.join(OUTPUT_FOLDER, output_filename)
#         wb.save(output_path)
#         logging.info(f"엑셀 파일이 성공적으로 저장되었습니다: {output_path}")
#         return output_path
#     except Exception as e:
#         logging.error(f"엑셀 파일 생성 중 오류 발생 (고객명: {client_name}, 배송일자: {order_date}): {e}")
#         raise e
def export_client_orders_to_files(from_date, to_date, client_code):
    logging.debug("export_client_orders_to_files 함수 시작 (거래처별 통합 PDF 병합)")
    logging.info(f"매개변수: from_date={from_date}, to_date={to_date}, client_code='{client_code}'")
    
    engine = get_sqlalchemy_engine()
    if engine is None:
        logging.error("SQLAlchemy 엔진 생성 실패")
        raise ConnectionError("SQLAlchemy 엔진 생성 실패")
    
    df = fetch_client_data(engine, from_date, to_date, client_code)
    logging.info(f"fetch_client_data 완료: {len(df)} 행 조회됨")
    if df.empty:
        logging.error("조회된 데이터가 없습니다.")
        raise ValueError("조회된 데이터가 없습니다.")
    
    try:
        df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce').dt.strftime('%Y-%m-%d')
        logging.debug("데이터 전처리 완료")
    except Exception as e:
        logging.error(f"데이터 전처리 오류: {e}", exc_info=True)
        raise e

    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
        logging.info(f"출력 폴더 생성: {OUTPUT_FOLDER}")
    
    merged_file_paths = []  # 각 거래처별 최종 merged pdf 파일 경로 저장

    # 거래처별 그룹화 (client_code 기준)
    grouped_clients = df.groupby(['client_code'])
    logging.info(f"기간 {from_date} ~ {to_date} 데이터: {len(grouped_clients)} 거래처 그룹")
    
    for client_code, client_group in grouped_clients:
        client_name = client_group['client_name'].iloc[0] if 'client_name' in client_group.columns else "unknown"
        logging.info(f"Processing client_code: {client_code}, client_name: {client_name}")
        
        daily_pdf_files = []  # 해당 거래처의 일자별 PDF 파일 목록 (tuple: (order_date, pdf_path))
        
        # 거래일자별 그룹화
        grouped_dates = client_group.groupby(['order_date'])
        for order_date, group in grouped_dates:
            logging.info(f"처리 일자: {order_date} (client: {client_code})")
            try:
                wb, ws = load_excel_template()
            except Exception as e:
                logging.error(f"엑셀 템플릿 로드 오류: {e}")
                continue
            try:
                # generate_excel_file() 함수는 개별 일자별 엑셀과 PDF를 생성함
                excel_path = generate_excel_file(wb, ws, client_name, order_date, group)
                logging.info(f"{order_date} 엑셀 파일 생성 완료: {excel_path}")
            except Exception as e:
                logging.error(f"엑셀 파일 생성 오류 (client_code: {client_code}, order_date: {order_date}): {e}", exc_info=True)
                continue
            try:
                pdf_path = convert_excel_to_pdf(excel_path, OUTPUT_FOLDER)
                daily_pdf_files.append((order_date, pdf_path))
                logging.info(f"{order_date} PDF 파일 생성 완료: {pdf_path}")
            except Exception as e:
                logging.error(f"PDF 변환 오류 (client_code: {client_code}, order_date: {order_date}): {e}", exc_info=True)
                continue
        
        if daily_pdf_files:
            # 정렬: 날짜 오름차순 정렬 (order_date가 "YYYY-MM-DD" 형식이므로 문자열 정렬 가능)
            daily_pdf_files.sort(key=lambda x: x[0])
            pdf_paths = [pdf for date, pdf in daily_pdf_files]
            final_filename = f"거래명세표_{client_name}_{from_date.replace('-', '')}_{to_date.replace('-', '')}.pdf"
            final_merged_pdf = os.path.join(OUTPUT_FOLDER, final_filename)
            try:
                merger = PdfMerger()
                for pdf in pdf_paths:
                    merger.append(pdf)
                merger.write(final_merged_pdf)
                merger.close()
                logging.info(f"최종 PDF 병합 완료 for client {client_code}: {final_merged_pdf}")
                merged_file_paths.append(final_merged_pdf)
            except Exception as e:
                logging.error(f"최종 PDF 병합 오류 for client {client_code}: {e}", exc_info=True)
                continue
        else:
            logging.error(f"생성된 PDF 파일이 없습니다 for client {client_code}.")
    
    if merged_file_paths:
        # 만약 한 거래처만 선택되었다면 리스트의 첫 번째 파일을 반환하거나,
        # 여러 거래처의 최종 PDF 파일 목록을 반환할 수 있습니다.
        # 여기서는 최종 PDF 파일 목록(거래처별)이 반환됩니다.
        return merged_file_paths
    else:
        logging.error("생성된 병합 PDF 파일이 없습니다.")
        raise ValueError("생성된 병합 PDF 파일이 없습니다.")


def generate_excel_file(wb, ws, client_name, order_date, group):
    try:
        # order_date가 tuple인 경우 첫 번째 요소를 사용
        if isinstance(order_date, tuple):
            order_date_str = order_date[0]
        else:
            order_date_str = order_date

        client_info = {
            'full_name': group['full_name'].iloc[0],
            'reg_no': group['reg_no'].iloc[0],
            'president': group['president'].iloc[0],
            'address1': group['address1'].iloc[0]
        }
        # wb를 첫번째 인자로 전달 (ws는 두번째)
        insert_data_to_excel(wb, ws, SUPPLIER_INFO, client_info, order_date_str, group.to_dict('records'))
        sanitized_client_name = re.sub(r'[\\/*?:"<>|]', "_", client_name)
        sanitized_order_date = order_date_str.replace('-', '')
        if sanitized_client_name == "":
            sanitized_client_name = "unknown"
        output_filename = f"거래명세표_{sanitized_client_name}_{sanitized_order_date}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)
        wb.save(output_path)
        logging.info(f"엑셀 파일이 성공적으로 저장되었습니다: {output_path}")
        return output_path
    except Exception as e:
        logging.error(f"엑셀 파일 생성 중 오류 발생 (고객명: {client_name}, 거래일자: {order_date}): {e}")
        raise e

def insert_data_to_excel(wb, ws, supplier_info, client_info, order_date, data_rows):
    ws['G3'] = supplier_info['등록번호']
    ws['G4'] = supplier_info['상호 (법인명)']
    ws['G5'] = supplier_info['성명']
    ws['G6'] = supplier_info['주소']
    reg_no = client_info.get('reg_no', '-').replace('-', '')  # 하이픈 제거
    reg_no = reg_no.ljust(10, '-')[:10]  # 10자리 맞추기
    reg_no_cells = ['V3', 'W3', 'X3', 'Z3', 'AA3', 'AC3', 'AD3', 'AE3', 'AF3', 'AG3']
    for cell, char in zip(reg_no_cells, reg_no):
        column_letters = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell)))
        column_num = column_index_from_string(column_letters)
        insert_cell_value(ws, row_num, column_num, char)
    ws['V4'] = f"{client_info.get('full_name', '-')}"
    ws['V5'] = f"{client_info.get('president', '-')}"
    ws['V6'] = f"{client_info.get('address1', '-')}"
    if data_rows:
        rep_name = data_rows[0].get('rep_name', '-')
        ws['U2'] = rep_name
    else:
        ws['U2'] = '-'
    ws['A2'] = f"{order_date}"
    current_row = 8
    total_order_amount = 0.0
    total_vat = 0.0
    for row in data_rows:
            try:
                # 각 행의 값을 float으로 변환
                qty = float(row['qty'])
                unit_price = float(row['unit_price'])
                order_amount = float(row['order_amount'])
                vat = float(row['vat'])
                total_amount = float(row['total_amount'])

                # 주문금액이 음수이면 양수로 변경 (부가세도 동일하게)
                if total_amount < 0:
                    order_amount *= -1
                    vat *= -1

                insert_cell_value(ws, current_row, 1, row['item_name'])
                insert_cell_value(ws, current_row, 14, row['unit'])
                insert_cell_value(ws, current_row, 17, qty)
                insert_cell_value(ws, current_row, 21, unit_price)
                insert_cell_value(ws, current_row, 26, order_amount)
                insert_cell_value(ws, current_row, 30, vat)

                total_order_amount += order_amount
                total_vat += vat
                current_row += 1
            except Exception as e:
                logging.error(f"데이터 삽입 중 오류 발생 (행: {current_row}): {e}")
                continue  # 다음 행으로 넘어가기
    current_row +=1
    try:
        ws['Z43'] = total_order_amount
        ws['AD43'] = total_vat
        ws['M44'] = total_order_amount + total_vat  # M43로 변경
        ws['Z43'].number_format = '#,##0'
        ws['AD43'].number_format = '#,##0'
        ws['M44'].number_format = '#,##0'
        logging.info(f"합계 - order_amount: {total_order_amount}, vat: {total_vat}, total: {total_order_amount + total_vat}이(가) Z43, AD43, M44 셀에 할당되었습니다.")
    except Exception as e:
        logging.error(f"합계 셀 할당 중 오류 발생: {e}")
    logging.info(f"엑셀에 {len(data_rows)}개의 데이터가 삽입되었습니다.")

def convert_excel_to_pdf(excel_path, output_folder):
    try:
        command = [
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_folder,
            excel_path
        ]
        logging.info(f"LibreOffice를 사용하여 PDF로 변환 중: {excel_path}")
        subprocess.run(command, check=True)
        base_filename = os.path.splitext(os.path.basename(excel_path))[0]
        pdf_filename = f"{base_filename}.pdf"
        pdf_path = os.path.join(output_folder, pdf_filename)
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF 변환이 실패했습니다: {pdf_path}")
        logging.info(f"PDF 파일이 성공적으로 생성되었습니다: {pdf_path}")
        return pdf_path
    except subprocess.CalledProcessError as e:
        logging.error(f"LibreOffice 변환 오류: {e}")
        raise e
    except Exception as e:
        logging.error(f"PDF 변환 중 오류 발생: {e}")
        raise e

def fetch_data(engine, target_date):
    if engine is None:
        logging.error("SQLAlchemy 엔진이 생성되지 않았습니다.")
        return None
    try:
        query = f"""
            SELECT 
                a.order_date,
                a.rep_code,
                a.rep_name,
                a.client_code,
                a.client_name,
                a.item_code,
                a.item_name,
                a.cond,
                a.unit,
                a.qty,
                a.cal_qty,
                a.unit_price,
                a.order_amount,
                a.vat,
                a.total_amount,
                a.tax,
                c.full_name,       -- 매출처명
                c.reg_no,          -- 매출처 등록번호
                c.president,       -- 매출처 성명
                c.address1         -- 매출처 주소
            FROM {AR_ORDER_DETAILS_ITEM_TABLE} a
            LEFT JOIN {CM_CHAIN_TABLE} c ON a.client_code COLLATE utf8mb4_unicode_ci = c.chain_no COLLATE utf8mb4_unicode_ci
            WHERE a.order_date = '{target_date}'
            ORDER BY a.client_code, a.order_date
        """
        df = pd.read_sql(query, engine)
        logging.info(f"MySQL에서 데이터를 성공적으로 조회했습니다. 총 {len(df)}개의 레코드.")
        return df
    except Exception as e:
        logging.error(f"데이터 조회 중 오류 발생: {e}")
        return None

def export_orders_to_files(order_date):
    setup_export_logging()
    engine = get_sqlalchemy_engine()
    if engine is None:
        logging.error("SQLAlchemy 엔진 생성 실패")
        raise ConnectionError("SQLAlchemy 엔진 생성 실패")
    df = fetch_data(engine, order_date)
    if df is None or df.empty:
        logging.error("데이터 조회 실패 또는 데이터 없음")
        raise ValueError("데이터 조회 실패 또는 데이터 없음")
    grouped = df.groupby(['client_code'])
    logging.info(f"배송일자 {order_date} 데이터: {len(grouped)} 거래처 그룹")
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
        logging.info(f"출력 폴더 생성: {OUTPUT_FOLDER}")
    file_paths = []
    for client_code, group in grouped:
        client_name = group['client_name'].iloc[0]
        logging.info(f"Processing client_code: {client_code}, client_name: {client_name}")
        try:
            df_processed = group.copy()
            df_processed = preprocess_data(df_processed)
        except Exception as e:
            logging.error(f"전처리 오류 (client_code: {client_code}): {e}")
            continue
        if pd.isna(client_name) or client_name.strip() == "":
            logging.error(f"클라이언트 이름 누락 (client_code: {client_code}). 건너뜀.")
            continue
        try:
            wb, ws = load_excel_template()
        except Exception as e:
            logging.error(f"엑셀 템플릿 로드 오류: {e}")
            continue
        try:
            excel_path = generate_excel_file(wb, ws, client_name, order_date, df_processed)
            file_paths.append(excel_path)
            logging.info(f"{order_date} 엑셀 파일 생성 완료: {excel_path}")
            try:
                pdf_path = convert_excel_to_pdf(excel_path, OUTPUT_FOLDER)
                file_paths.append(pdf_path)
                logging.info(f"{order_date} PDF 파일 생성 완료: {pdf_path}")
            except Exception as e:
                logging.error(f"PDF 변환 오류 (client_code: {client_code}, order_date: {order_date}): {e}")
        except Exception as e:
            logging.error(f"엑셀 파일 생성 오류 (client_code: {client_code}, order_date: {order_date}): {e}")
    logging.info("모든 개별 PDF 파일 생성 완료. 병합 시작.")
    if file_paths:
        merged_pdf_path = os.path.join(OUTPUT_FOLDER, f"merged_orders_{order_date}.pdf")
        try:
            merger = PdfMerger()
            # file_paths 리스트에서 PDF 파일만 선택 (확장자가 .pdf)
            pdf_files = [fp for fp in file_paths if fp.lower().endswith('.pdf')]
            for pdf in pdf_files:
                merger.append(pdf)
            merger.write(merged_pdf_path)
            merger.close()
            logging.info(f"PDF 병합 완료: {merged_pdf_path}")
            return merged_pdf_path
        except Exception as e:
            logging.error(f"PDF 병합 오류: {e}", exc_info=True)
            raise e
    else:
        logging.error("생성된 PDF 파일이 없습니다.")
        raise ValueError("생성된 PDF 파일이 없습니다.")
    
def setup_export_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("export_orders.log"),
            logging.StreamHandler()
        ]
    )
    logging.info("내보내기 프로세스 시작.")

def get_sqlalchemy_engine():
    try:
        engine = create_engine(
            f"mysql+pymysql://{os.getenv('DB_USER', 'nolboo')}:{os.getenv('DB_PASSWORD', '2024!puser')}@"
            f"{os.getenv('DB_HOST', '175.196.7.45')}/{os.getenv('DB_NAME', 'nolboo')}?charset=utf8mb4"
        )
        logging.info("SQLAlchemy 엔진이 성공적으로 생성되었습니다.")
        return engine
    except Exception as e:
        logging.error(f"SQLAlchemy 엔진 생성 오류: {e}")
        return None

def preprocess_data(df):
    required_columns = list(COLUMN_MAPPING.values()) + ['full_name', 'reg_no', 'president', 'address1']  # 추가 컬럼
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.error(f"MySQL 데이터에 누락된 컬럼이 있습니다: {missing_columns}")
        raise ValueError(f"MySQL 데이터에 누락된 컬럼이 있습니다: {missing_columns}")
    df = df[required_columns]
    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce').dt.strftime('%Y-%m-%d')
    decimal_columns = ['qty', 'cal_qty', 'unit_price', 'order_amount', 'vat', 'total_amount']
    for col in decimal_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).round(2)
    string_fields = ['rep_code', 'rep_name', 'client_code', 'client_name', 'item_code', 'item_name', 'cond', 'unit', 'tax']
    for field in string_fields:
        df[field] = df[field].astype(str).str.strip()
    df['full_name'] = df['full_name'].fillna('-').astype(str).str.strip()
    df['reg_no'] = df['reg_no'].fillna('-').astype(str).str.strip()
    df['president'] = df['president'].fillna('-').astype(str).str.strip()
    df['address1'] = df['address1'].fillna('-').astype(str).str.strip()
    logging.info("데이터 전처리 및 형변환이 완료되었습니다.")
    return df

# ------------------------
# 백그라운드 작업 관련 함수 및 전역 변수
# ------------------------
background_tasks = {}  # task_id -> {'status': 'pending'|'complete'|'failed', 'result': 파일 경로 목록 or error message}

def background_export_client_orders(from_date, to_date, client_code, task_id):
    try:
        logging.info(f"백그라운드 작업 시작: task_id={task_id}")
        file_paths = export_client_orders_to_files(from_date, to_date, client_code)
        update_task_status(task_id, 'complete', json.dumps(file_paths))
        logging.info(f"백그라운드 작업 완료: task_id={task_id}")
    except Exception as e:
        logging.error(f"백그라운드 작업 오류 (task_id={task_id}): {e}", exc_info=True)
        update_task_status(task_id, 'failed', str(e))

# ---------------------------
# 신규 라우트: 매출처/일자 범위 거래명세표 다운로드 폼 (드롭다운 + 검색 기능)
# ---------------------------
@app.route('/download_client_orders_form', methods=['GET'])
def download_client_orders_form():
    form = DownloadClientOrdersForm()
    db = get_db_connection()
    if db:
        try:
            with db.cursor(dictionary=True) as cursor:
                cursor.execute("""
                    SELECT DISTINCT c.chain_no, c.full_name 
                    FROM cm_chain c
                    INNER JOIN AROrderDetailsItem a ON a.client_code = c.chain_no
                    ORDER BY c.full_name
                """)
                clients = cursor.fetchall()
                form.client_code.choices = [('', '전체매출처')] + [
                    (client['chain_no'], f"{client['chain_no']} - {client['full_name']}")
                    for client in clients
                ]
        except Exception as e:
            logging.error(f"cm_chain 조회 오류: {e}", exc_info=True)
            form.client_code.choices = [('', '전체매출처')]
        finally:
            db.close()
    else:
        form.client_code.choices = [('', '전체매출처')]
    return render_template('download_client_orders_form.html', form=form)

# ------------------------
# 라우트: 매출처/일자 범위 거래명세표 생성 요청 (백그라운드 처리)
# ------------------------
# @app.route('/download_client_orders', methods=['POST'])
# def download_client_orders():
#     form = DownloadClientOrdersForm()
#     # POST 요청 시 client_code 선택지를 재설정
#     db = get_db_connection()
#     if db:
#         try:
#             with db.cursor(dictionary=True) as cursor:
#                 cursor.execute("SELECT chain_no, full_name FROM cm_chain ORDER BY full_name")
#                 clients = cursor.fetchall()
#                 form.client_code.choices = [('', '전체매출처')] + [
#                     (client['chain_no'], f"{client['chain_no']} - {client['full_name']}")
#                     for client in clients
#                 ]
#             logging.info(f"POST: client_code 선택지 재설정 완료. {form.client_code.choices}")
#         except Exception as e:
#             logging.error(f"cm_chain 조회 오류 (POST): {e}", exc_info=True)
#             form.client_code.choices = [('', '전체매출처')]
#         finally:
#             db.close()
#     else:
#         form.client_code.choices = [('', '전체매출처')]

#     if form.validate_on_submit():
#         client_code = form.client_code.data.strip() if form.client_code.data else ""
#         from_date = form.from_date.data.strftime('%Y-%m-%d')
#         to_date = form.to_date.data.strftime('%Y-%m-%d')
#         logging.info(f"매출처 거래명세표 생성 요청 - client_code: '{client_code}', 기간: {from_date} ~ {to_date}")
#         try:
#             task_id = str(uuid.uuid4())
#             insert_task(task_id, 'pending')
#             thread = threading.Thread(target=background_export_client_orders, args=(from_date, to_date, client_code, task_id))
#             thread.start()
#             flash("파일 생성 작업이 백그라운드에서 시작되었습니다. 작업 완료 후 다운로드 페이지에서 확인하세요.", "info")
#             return redirect(url_for('download_client_orders_status', task_id=task_id))
#         except Exception as e:
#             logging.error(f"매출처 거래명세표 생성 오류: {e}", exc_info=True)
#             flash(f"매출처 거래명세표 생성 중 오류가 발생했습니다: {e}", 'danger')
#             return redirect(url_for('download_client_orders_form'))
#     else:
#         logging.error(f"DownloadClientOrdersForm: 폼 검증 실패 - {form.errors}")
#         for field, errors in form.errors.items():
#             for error in errors:
#                 flash(f"{getattr(form, field).label.text} - {error}", 'danger')
#         return redirect(url_for('download_client_orders_form'))

# ------------------------
# 라우트: 백그라운드 작업 상태 및 다운로드 페이지
# ------------------------
# @app.route('/download_client_orders_status', methods=['GET'])
# def download_client_orders_status():
#     task_id = request.args.get('task_id', None)
#     if not task_id:
#         flash("작업 ID가 제공되지 않았습니다.", "danger")
#         return redirect(url_for('download_client_orders_form'))
#     task = get_task(task_id)
#     if not task:
#         flash("유효하지 않은 작업 ID입니다.", "danger")
#         return redirect(url_for('download_client_orders_form'))
#     if task['status'] == 'pending':
#         return render_template('download_client_orders_status.html', status="진행중", task_id=task_id)
#     elif task['status'] == 'failed':
#         flash(f"작업 실패: {task['result']}", "danger")
#         return redirect(url_for('download_client_orders_form'))
#     elif task['status'] == 'complete':
#         # 결과로 저장된 파일 경로 목록(JSON 문자열)을 복원
#         try:
#             file_paths = json.loads(task['result'])
#         except Exception as e:
#             logging.error(f"작업 결과 파싱 오류: {e}", exc_info=True)
#             flash("작업 결과를 처리하는 중 오류가 발생했습니다.", "danger")
#             return redirect(url_for('download_client_orders_form'))
#         # ZIP 파일 생성 후 다운로드
#         if file_paths:
#             with tempfile.TemporaryDirectory() as tmpdirname:
#                 zip_filename = f"거래명세표_{task_id}.zip"
#                 zip_path = os.path.join(tmpdirname, zip_filename)
#                 with ZipFile(zip_path, 'w') as zipf:
#                     for file_path in file_paths:
#                         zipf.write(file_path, os.path.basename(file_path))
#                 return send_from_directory(
#                     directory=tmpdirname,
#                     path=zip_filename,
#                     as_attachment=True,
#                     download_name=zip_filename
#                 )
#         else:
#             flash("생성된 파일이 없습니다.", "danger")
#             return redirect(url_for('download_client_orders_form'))
#     else:
#         flash("알 수 없는 작업 상태입니다.", "danger")
#         return redirect(url_for('download_client_orders_form'))
# ------------------------
# Background task function
# ------------------------
def background_export_client_orders(from_date, to_date, client_code, task_id):
    try:
        logging.info(f"백그라운드 작업 시작: task_id={task_id}")
        # This function should merge the individual PDF files and return the final merged PDF path.
        final_pdf = export_client_orders_to_files(from_date, to_date, client_code)
        # Always store a list of file paths (even if one file)
        update_task_status(task_id, 'complete', json.dumps([final_pdf]))
        logging.info(f"백그라운드 작업 완료: task_id={task_id}")
    except Exception as e:
        logging.error(f"백그라운드 작업 오류 (task_id={task_id}): {e}", exc_info=True)
        update_task_status(task_id, 'failed', str(e))

# ------------------------
# Route: 요청을 받으면 start background task and show status page
# ------------------------
@app.route('/download_client_orders', methods=['POST'])
def download_client_orders():
    form = DownloadClientOrdersForm()
    # Reset choices on POST
    db = get_db_connection()
    if db:
        try:
            with db.cursor(dictionary=True) as cursor:
                cursor.execute("SELECT chain_no, full_name FROM cm_chain ORDER BY full_name")
                clients = cursor.fetchall()
                form.client_code.choices = [('', '전체매출처')] + [
                    (client['chain_no'], f"{client['chain_no']} - {client['full_name']}")
                    for client in clients
                ]
            logging.info(f"POST: client_code 선택지 재설정 완료. Choices: {form.client_code.choices}")
        except Exception as e:
            logging.error(f"cm_chain 조회 오류 (POST): {e}", exc_info=True)
            form.client_code.choices = [('', '전체매출처')]
        finally:
            db.close()
    else:
        form.client_code.choices = [('', '전체매출처')]

    if form.validate_on_submit():
        client_code = form.client_code.data.strip() if form.client_code.data else ""
        from_date = form.from_date.data.strftime('%Y-%m-%d')
        to_date = form.to_date.data.strftime('%Y-%m-%d')
        logging.info(f"매출처 거래명세표 생성 요청 - client_code: '{client_code}', 기간: {from_date} ~ {to_date}")
        try:
            task_id = str(uuid.uuid4())
            insert_task(task_id, 'pending')
            thread = threading.Thread(target=background_export_client_orders,
                                      args=(from_date, to_date, client_code, task_id))
            thread.start()
            flash("파일 생성 작업이 백그라운드에서 시작되었습니다. 잠시 후 '다운로드 상태' 페이지를 새로고침해 주세요.", "info")
            # Instead of redirecting immediately, render a status page
            return render_template('download_client_orders_status.html', task_id=task_id)
        except Exception as e:
            logging.error(f"매출처 거래명세표 생성 오류: {e}", exc_info=True)
            flash(f"매출처 거래명세표 생성 중 오류가 발생했습니다: {e}", 'danger')
            return redirect(url_for('download_client_orders_form'))
    else:
        logging.error(f"DownloadClientOrdersForm: 폼 검증 실패 - {form.errors}")
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text} - {error}", 'danger')
        return redirect(url_for('download_client_orders_form'))


@app.route('/api/task_status', methods=['GET'])
def api_task_status():
    task_id = request.args.get('task_id')
    if not task_id:
        return jsonify({'error': 'No task_id provided'}), 400
    task = get_task(task_id)
    if not task:
        return jsonify({'error': 'Invalid task_id'}), 404
    # Return only the relevant fields.
    return jsonify({
        'task_id': task_id,
        'status': task.get('status', ''),
        'result': task.get('result', '')
    })

# ------------------------
# Route: Status page (polled or auto-refreshed) and final download
# ------------------------
def merge_client_pdfs(from_date, to_date):
    """
    OUTPUT_FOLDER 내의 파일 중 파일명이
    "거래명세표_{client_name}_{YYYYMMDD}.pdf"인 파일들을
    거래처별로 그룹화하여, from_date ~ to_date 범위 내의 파일을 병합하고
    최종 파일명을 "거래명세표_{client_name}_{from_date(YYYYMMDD)}_{to_date(YYYYMMDD)}.pdf"로 생성한다.
    
    Returns:
         dict: {client_name: final_merged_pdf_path, ...}
    """
    merged_dict = {}
    # from_date, to_date: "YYYY-MM-DD" 형식 → 날짜 객체로 변환
    start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(to_date, "%Y-%m-%d").date()
    
    # OUTPUT_FOLDER 내의 모든 PDF 파일 목록을 가져옵니다.
    all_files = os.listdir(OUTPUT_FOLDER)
    pdf_files = [f for f in all_files if f.lower().endswith('.pdf')]
    
    # 파일명 형식: "거래명세표_{client_name}_{YYYYMMDD}.pdf"
    pattern = r"거래명세표_(.+?)_(\d{8})\.pdf"
    client_files = {}
    for filename in pdf_files:
        match = re.match(pattern, filename)
        if match:
            client_name = match.group(1)
            file_date_str = match.group(2)  # 예: "20250301"
            try:
                file_date = datetime.strptime(file_date_str, "%Y%m%d").date()
            except Exception as e:
                logging.error(f"파일 날짜 변환 오류: {filename}, {e}")
                continue
            # 기간 내 파일만 사용
            if start_date <= file_date <= end_date:
                client_files.setdefault(client_name, []).append(os.path.join(OUTPUT_FOLDER, filename))
    
    # 각 거래처별로 PDF 병합
    for client_name, files in client_files.items():
        # 정렬 (파일명이 날짜순이 되도록 정렬)
        files.sort()
        final_filename = f"거래명세표_{client_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
        final_path = os.path.join(OUTPUT_FOLDER, final_filename)
        try:
            merger = PdfMerger()
            for pdf in files:
                merger.append(pdf)
            merger.write(final_path)
            merger.close()
            logging.info(f"클라이언트 {client_name} 최종 병합 PDF 생성 완료: {final_path}")
            merged_dict[client_name] = final_path
        except Exception as e:
            logging.error(f"클라이언트 {client_name} PDF 병합 오류: {e}", exc_info=True)
    
    return merged_dict

# 비동기로 다운로드 잘되는 코드
# @app.route('/download_client_orders_status', methods=['GET'])
# def download_client_orders_status():
#     task_id = request.args.get('task_id', None)
#     if not task_id:
#         flash("작업 ID가 제공되지 않았습니다.", "danger")
#         return redirect(url_for('download_client_orders_form'))
#     task = get_task(task_id)
#     if not task:
#         flash("유효하지 않은 작업 ID입니다.", "danger")
#         return redirect(url_for('download_client_orders_form'))
#     if task['status'] == 'pending':
#         # Render a status page with JavaScript to auto-refresh every few seconds.
#         return render_template('download_client_orders_status.html', status="진행중", task_id=task_id)
#     elif task['status'] == 'failed':
#         flash(f"작업 실패: {task['result']}", "danger")
#         return redirect(url_for('download_client_orders_form'))
#     elif task['status'] == 'complete':
#         try:
#             file_paths = json.loads(task['result'])
#         except Exception as e:
#             logging.error(f"작업 결과 파싱 오류: {e}", exc_info=True)
#             flash("작업 결과를 처리하는 중 오류가 발생했습니다.", "danger")
#             return redirect(url_for('download_client_orders_form'))
#         if file_paths and isinstance(file_paths, list) and len(file_paths) > 0:
#             zip_filename = f"거래명세표_{task_id}.zip"
#             zip_path = os.path.join(OUTPUT_FOLDER, zip_filename)
#             try:
#                 with ZipFile(zip_path, 'w') as zipf:
#                     for file_path in file_paths:
#                         if os.path.isfile(file_path):
#                             zipf.write(file_path, os.path.basename(file_path))
#                         else:
#                             logging.error(f"파일 경로가 파일이 아님: {file_path}")
#                 logging.info(f"ZIP 파일 생성 완료: {zip_path}")
#                 return send_file(zip_path, as_attachment=True, download_name=zip_filename)
#             except Exception as e:
#                 logging.error(f"ZIP 파일 생성 오류: {e}", exc_info=True)
#                 flash("파일 압축 중 오류가 발생했습니다.", "danger")
#                 return redirect(url_for('download_client_orders_form'))
#         else:
#             flash("생성된 파일이 없습니다.", "danger")
#             return redirect(url_for('download_client_orders_form'))
#     else:
#         flash("알 수 없는 작업 상태입니다.", "danger")
#         return redirect(url_for('download_client_orders_form'))

@app.route('/download_client_orders_file', methods=['GET'])
def download_client_orders_file():
    task_id = request.args.get('task_id', None)
    if not task_id:
        flash("작업 ID가 제공되지 않았습니다.", "danger")
        return redirect(url_for('download_client_orders_form'))
    
    task = get_task(task_id)
    if not task:
        flash("유효하지 않은 작업 ID입니다.", "danger")
        return redirect(url_for('download_client_orders_form'))
    
    if task['status'] != 'complete':
        flash("작업이 아직 완료되지 않았습니다.", "warning")
        return redirect(url_for('download_client_orders_status', task_id=task_id))
    
    try:
        result = json.loads(task['result'])
    except Exception as e:
        logging.error(f"작업 결과 파싱 오류: {e}", exc_info=True)
        flash("작업 결과를 처리하는 중 오류가 발생했습니다.", "danger")
        return redirect(url_for('download_client_orders_form'))
    
    # result가 리스트가 아닌 경우 리스트로 변환
    if not isinstance(result, list):
        result = [result]
    
    # 중첩 리스트 평탄화 (이미 정의한 flatten_list 함수 사용)
    pdf_files = [fp for fp in flatten_list(result)
                 if isinstance(fp, str) and os.path.isfile(fp) and fp.lower().endswith('.pdf')]
    
    if not pdf_files:
        flash("생성된 PDF 파일이 없습니다.", "danger")
        return redirect(url_for('download_client_orders_form'))
    
    # 모든 PDF 파일을 ZIP 파일로 압축
    with tempfile.TemporaryDirectory() as tmpdirname:
        zip_filename = f"거래명세표_{task_id}.zip"
        zip_path = os.path.join(tmpdirname, zip_filename)
        try:
            with ZipFile(zip_path, 'w') as zipf:
                for pdf in pdf_files:
                    zipf.write(pdf, os.path.basename(pdf))
            logging.info(f"ZIP 파일 생성 완료: {zip_path}")
            response = send_file(zip_path, as_attachment=True, download_name=zip_filename)
            # 다운로드 후 5초 뒤 download_client_orders_form 화면으로 이동
            response.headers["Refresh"] = "5; url=" + url_for("download_client_orders_form")
            return response
        except Exception as e:
            logging.error(f"ZIP 파일 생성 오류: {e}", exc_info=True)
            flash("파일 압축 중 오류가 발생했습니다.", "danger")
            return redirect(url_for('download_client_orders_form'))

def flatten_list(lst):
    """중첩 리스트를 평탄화하는 함수"""
    flat = []
    for item in lst:
        if isinstance(item, list):
            flat.extend(flatten_list(item))
        else:
            flat.append(item)
    return flat
@app.route('/download_client_orders_status', methods=['GET'])
def download_client_orders_status():
    task_id = request.args.get('task_id', None)
    if not task_id:
        flash("작업 ID가 제공되지 않았습니다.", "danger")
        return redirect(url_for('download_client_orders_form'))
    task = get_task(task_id)
    if not task:
        flash("유효하지 않은 작업 ID입니다.", "danger")
        return redirect(url_for('download_client_orders_form'))
    
    if task['status'] == 'pending':
        return render_template('download_client_orders_status.html', status="진행중", task_id=task_id)
    elif task['status'] == 'failed':
        flash(f"작업 실패: {task['result']}", "danger")
        return redirect(url_for('download_client_orders_form'))
    elif task['status'] == 'complete':
        try:
            file_paths = json.loads(task['result'])
        except Exception as e:
            logging.error(f"작업 결과 파싱 오류: {e}", exc_info=True)
            flash("작업 결과를 처리하는 중 오류가 발생했습니다.", "danger")
            return redirect(url_for('download_client_orders_form'))
        
        # 중첩된 file_paths를 평탄화
        flat_file_paths = flatten_list(file_paths)
        # PDF 파일 경로만 선택 (문자열이며 실제 파일이 존재하고 확장자가 .pdf인 경우)
        pdf_files = [fp for fp in flat_file_paths if isinstance(fp, str) and os.path.isfile(fp) and fp.lower().endswith('.pdf')]
        
        if pdf_files:
            zip_filename = f"거래명세표_{task_id}.zip"
            with tempfile.TemporaryDirectory() as tmpdirname:
                zip_path = os.path.join(tmpdirname, zip_filename)
                try:
                    with ZipFile(zip_path, 'w') as zipf:
                        for fp in pdf_files:
                            zipf.write(fp, os.path.basename(fp))
                    logging.info(f"ZIP 파일 생성 완료: {zip_path}")
                    return send_file(zip_path, as_attachment=True, download_name=zip_filename)
                except Exception as e:
                    logging.error(f"ZIP 파일 생성 오류: {e}", exc_info=True)
                    flash("파일 압축 중 오류가 발생했습니다.", "danger")
                    return redirect(url_for('download_client_orders_form'))
        else:
            flash("생성된 PDF 파일이 없습니다.", "danger")
            return redirect(url_for('download_client_orders_form'))
    else:
        flash("알 수 없는 작업 상태입니다.", "danger")
        return redirect(url_for('download_client_orders_form'))

def etl_process():
    # excel_path_step3 = None  # 초기화
    try:
        logging.info("ETL 프로세스 시작.")

        # 현재 시간까지 포함한 고유한 파일명 생성
        now = datetime.now()
        today_str = now.strftime("%Y%m%d")
        timestamp = now.strftime("%H%M%S")
        sale_dy = now.strftime("%Y-%m-%d")  # 엑셀의 DATE 형식에 맞게 변환

        # Informix 연결 설정
        logging.info(f"Informix 연결 중: {informix_hostname}:{informix_port}")
        informix_conn = jaydebeapi.connect(
            informix_jdbc_driver_class,
            informix_jdbc_url,
            [informix_username, informix_password],
            jdbc_driver_path
        )
        informix_cursor = informix_conn.cursor()
        logging.info("Informix 연결 성공.")


        # 2단계 전 Delete
        try:
            # ----------------------------
            # 레코드 존재 여부 확인
            # ----------------------------
            count_query = "SELECT COUNT(*) FROM t_po_order_master WHERE date = ?"
            logging.info(f"오늘자 t_po_order_master 레코드 수 확인 SQL: {count_query} with date: {today_str}")
            informix_cursor.execute(count_query, (today_str,))
            count_result = informix_cursor.fetchall()
            record_count = count_result[0][0] if count_result else 0
            logging.info(f"오늘자 t_po_order_master 레코드 수: {record_count}")

            if record_count > 0:
                # ----------------------------
                # 레코드가 존재하므로 DELETE 실행
                # ----------------------------
                delete_query = "DELETE FROM t_po_order_master WHERE date = ?"
                logging.info(f"오늘자 t_po_order_master 레코드 삭제 SQL: {delete_query} with date: {today_str}")
                informix_cursor.execute(delete_query, (today_str,))
                # informix_conn.commit()  # 변경 사항을 커밋
                logging.info("오늘자 t_po_order_master 레코드 삭제 완료.")
                # ----------------------------
                # DELETE 쿼리 실행 끝
                # ----------------------------
            else:
                logging.info("오늘자 t_po_order_master 레코드가 존재하지 않으므로 DELETE를 건너뜁니다.")
            # ----------------------------
            # 레코드 존재 여부 확인 끝
            # ----------------------------
        except jaydebeapi.DatabaseError as db_err:
            logging.exception("DELETE 쿼리 실행 중 Database 에러 발생:")
            informix_conn.rollback()  # 롤백
            raise  # 원래의 예외를 다시 발생시킴
        except Exception as e:
            logging.exception("DELETE 쿼리 실행 중 예외 발생:")
            informix_conn.rollback()  # 롤백
            raise  # 원래의 예외를 다시 발생시킴


        # 2단계: pr_order_data_load 프로시저 호출 (CALL 방식)
        logging.info("2단계: pr_order_data_load 프로시저 호출 (CALL 방식)")
        p_ord_date = today_str  # 파라미터로 오늘 날짜 사용
        p_proc_fg = '0'  # 프로시저 실행 플래그 (필요에 따라 변경)

        # Informix에서 프로시저 호출 (CALL 구문 사용)
        proc_call = f"CALL pr_order_data_load('{p_ord_date}', '{p_proc_fg}')"
        logging.info(f"프로시저 호출 SQL: {proc_call}")
        informix_cursor.execute(proc_call)

        # 프로시저 반환값을 가져오기 위해 fetchall을 사용
        proc_result = informix_cursor.fetchall()

        # 프로시저가 반환한 값의 수를 확인
        expected_return_count = 5  # r_rtn_code, r_rtn_desc, r_pos_cnt, r_web_cnt, r_ars_cnt
        if not proc_result:
            logging.error("프로시저 반환값이 없습니다.")
            raise ValueError("프로시저 반환값이 없습니다.")
        elif len(proc_result[0]) < expected_return_count:
            logging.error(f"프로시저 반환값의 수가 예상과 다릅니다. 예상: {expected_return_count}, 실제: {len(proc_result[0])}")
            raise ValueError("프로시저 반환값의 수가 예상과 다릅니다.")
        else:
            # 반환된 값을 인덱스로 접근
            r_rtn_code, r_rtn_desc, r_pos_cnt, r_web_cnt, r_ars_cnt = proc_result[0]
            logging.info(f"프로시저 반환값 - 코드: {r_rtn_code}, 설명: {r_rtn_desc}, POS 건수: {r_pos_cnt}, WEB 건수: {r_web_cnt}, ARS 건수: {r_ars_cnt}")

            # 반환 코드에 따른 추가 로직 구현 가능
            if r_rtn_code == '1':
                logging.warning("처리할 데이터가 없습니다.")
            elif r_rtn_code == '2':
                logging.warning("이미 처리 완료 되었습니다.")
            elif r_rtn_code == '0':
                logging.info("정상 처리 완료.")

        # 3단계: 최종 데이터 추출
        logging.info("3단계: 최종 데이터 추출")
        query_step3 = f"""
        SELECT date, 
               full_name, 
               rechain_no, 
               rep_full_name, 
               item_no, 
               item_full_name, 
               qty, 
               time, 
               remark, 
               out_date, 
               item_price, 
               item_tax, 
               tax,
               (qty * (item_price + item_tax)) AS total
        FROM (
            SELECT a.date AS date, 
                   b.full_name AS full_name, 
                   b.rechain_no AS rechain_no, 
                   c.full_name AS rep_full_name, 
                   a.item_no AS item_no, 
                   d.full_name AS item_full_name, 
                   a.qty AS qty, 
                   a.time AS time, 
                   a.remark AS remark, 
                   a.out_date AS out_date, 
                   CASE 
                       WHEN b.contract_no = '2' THEN 
                           CASE 
                               WHEN d.PACKAGE_MODEL_PRICE = 0 THEN d.MODEL_PRICE 
                               ELSE d.PACKAGE_MODEL_PRICE 
                           END 
                       ELSE 
                           CASE 
                               WHEN d.PACKAGE_CHAIN_PRICE = 0 THEN d.CHAIN_PRICE 
                               ELSE d.PACKAGE_CHAIN_PRICE 
                           END 
                   END AS item_price,
                   CASE 
                       WHEN b.contract_no = '2' THEN 
                           CASE 
                               WHEN d.PACKAGE_MODEL_TAX = 0 THEN d.MODEL_TAX 
                               ELSE d.PACKAGE_MODEL_TAX 
                           END 
                       ELSE 
                           CASE 
                               WHEN d.PACKAGE_CHAIN_TAX = 0 THEN d.CHAIN_TAX 
                               ELSE d.PACKAGE_CHAIN_TAX 
                           END 
                   END AS item_tax,
                   CASE 
                        WHEN tax_type = '1' THEN 'Tax' 
                        ELSE 'No Tax' 
                   END AS tax
            FROM t_po_order_master AS a
            INNER JOIN cm_chain AS b ON a.chain_no = b.chain_no  
            INNER JOIN cm_chain AS c ON b.rechain_no = c.chain_no 
            INNER JOIN v_item_master AS d ON a.item_no = d.item_no 
            WHERE a.date ='{today_str}'
        ) subquery;
        """

        # Log the query string
        log_query_string(query_step3)

        try:
            df_step3 = extract_data(informix_cursor, query_step3)
            logging.info(f"3단계 데이터 추출 완료. 총 {len(df_step3)}개의 레코드.")
        except jaydebeapi.DatabaseError as db_err:
            logging.error(f"3단계 데이터 추출 중 오류 발생: {db_err}")
            logging.error(traceback.format_exc())
            raise db_err

        if df_step3.empty:
            logging.warning("추출된 데이터가 없습니다.")
        else:
            # 데이터 검증: 특수 문자 확인
            columns_to_convert = ['full_name', 'rep_full_name', 'item_full_name']
            check_special_characters(df_step3, columns_to_convert)

            # 'full_name', 'rep_full_name', 'item_full_name' 컬럼에 인코딩 변환 적용
            for col in columns_to_convert:
                if col in df_step3.columns:
                    df_step3[col] = df_step3[col].apply(convert_to_utf8)
                    logging.info(f"'{col}' 컬럼의 인코딩 변환 완료.")
                else:
                    logging.warning(f"'{col}' 컬럼이 데이터프레임에 존재하지 않습니다.")

            # 웹발주 데이터를 엑셀로 저장
            excel_output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_output')
            os.makedirs(excel_output_dir, exist_ok=True)
            excel_filename = f't_po_order_master_{today_str}_{timestamp}.xlsx'  # 고유한 파일명
            excel_path_step3 = os.path.join(excel_output_dir, excel_filename)
            save_to_excel(df_step3, excel_path_step3)
            #  여기에 MySQL insert 로직 추가

        # 연결 종료
        informix_cursor.close()
        informix_conn.close()
        logging.info("Informix 연결 종료.")

        logging.info("ETL 프로세스 성공적으로 완료.")

        return excel_path_step3

    except jaydebeapi.DatabaseError as db_err:
        logging.error(f"Database 에러 발생: {db_err}")
        logging.error(traceback.format_exc())
        raise db_err
    except Exception as e:
        logging.error(f"ETL 프로세스 실패: {e}")
        logging.error(traceback.format_exc())
        raise e  # 예외를 상위로 전달하여 Flask에서 처리하도록 함
    
    # ------------------------
    # 7. Flask 라우트 정의
    # ------------------------

    # 메인 페이지
@app.route('/')
def index():
    return render_template('index.html')

# 발주 내역 추가
@app.route('/add_order', methods=['GET', 'POST'])
def add_order():
    form = AddOrderForm()
    
    db = get_db_connection()
    if not db:
        flash('Database connection failed.', 'danger')
        return render_template('add_order.html', form=form)

    try:
        with db.cursor(dictionary=True) as cursor:
            # 거래처 목록을 동적으로 가져와 폼의 선택지에 추가
            try:
                cursor.execute("SELECT client_code, client_name FROM ARClientMaster")
                clients = cursor.fetchall()
                form.client_code.choices = [(client['client_code'], client['client_name']) for client in clients]
            except mysql.connector.Error as err:
                logging.error(f"거래처 목록 조회 실패: {err}")
                flash('거래처 목록을 불러오는 중 오류가 발생했습니다.', 'danger')
                form.client_code.choices = []

            if form.validate_on_submit():
                client_code = form.client_code.data
                order_date = form.order_date.data
                amount = form.amount.data

                logging.debug(f"발주 추가 요청 - client_code: {client_code}, order_date: {order_date}, amount: {amount}")

                try:
                    # 거래처 마스터에서 필요한 정보 조회 (client_code 기준, 첫 번째 레코드)
                    cursor.execute("SELECT representative_code, manager, client_name FROM ARClientMaster WHERE client_code = %s LIMIT 1", (client_code,))
                    client = cursor.fetchone()

                    if not client:
                        logging.warning(f"선택된 client_code '{client_code}'가 ARClientMaster 테이블에 존재하지 않습니다.")
                        flash('선택한 거래처가 존재하지 않습니다.', 'danger')
                        return render_template('add_order.html', form=form)  # form 전달

                    representative_code = client['representative_code']
                    manager = client['manager']
                    client_name = client['client_name']

                    logging.debug(f"대표 코드: {representative_code}, 관리자: {manager}")

                    # AROrderDetails 테이블에 삽입
                    insert_order_query = """
                        INSERT INTO AROrderDetails (
                            representative_code, client_code, client_name, collector_key, manager, order_date, order_amount
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s
                        )
                    """
                    cursor.execute(insert_order_query, (representative_code, client_code, client_name, '', manager, order_date, amount))

                    # ARTransactionsLedger 테이블에 식자재 매출 기록 삽입
                    insert_ledger_query = """
                        INSERT INTO ARTransactionsLedger (
                            transaction_date,
                            representative_code,
                            client,
                            outlet_name,
                            debit,
                            credit,
                            food_material_sales
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s
                        )
                    """
                    cursor.execute(insert_ledger_query, (
                        order_date,
                        client_code,  # client_code 사용
                        client_name,
                        '',  # outlet_name은 발주와 관련 없으므로 빈 문자열
                        amount,  # debit (식자재 매출)
                        0,       # credit
                        amount   # food_material_sales
                    ))

                    db.commit()

                    logging.info(f"발주 내역이 성공적으로 추가되었습니다: client_code={client_code}, order_date={order_date}, amount={amount}")
                    flash('발주 내역이 성공적으로 추가되었습니다.', 'success')
                    return redirect(url_for('index'))

                except mysql.connector.Error as db_err:
                    logging.error(f"발주 내역 추가 실패 (DB 오류): {db_err}")
                    db.rollback()
                    flash('발주 내역을 추가하는 중 오류가 발생했습니다.', 'danger')
                    return render_template('add_order.html', form=form)  # form 전달
                except Exception as e:
                    logging.error(f"발주 내역 추가 실패 (기타 오류): {e}")
                    db.rollback()
                    flash('발주 내역을 추가하는 중 예상치 못한 오류가 발생했습니다.', 'danger')
                    return render_template('add_order.html', form=form)  # form 전달

    finally:
        db.close()

    return render_template('add_order.html', form=form)

# 대표 코드 반환 API 엔드포인트
@app.route('/get_representative_code', methods=['POST'])
def get_representative_code():
    client_code = request.form.get('client_code')
    if not client_code:
        return jsonify({'error': 'Client code not provided.'}), 400

    db = get_db_connection()
    if not db:
        return jsonify({'error': 'Database connection failed.'}), 500

    try:
        with db.cursor(dictionary=True) as cursor:
            cursor.execute("SELECT representative_code FROM ARClientMaster WHERE client_code = %s LIMIT 1", (client_code,))
            result = cursor.fetchone()
            if result:
                return jsonify({'representative_code': result['representative_code']})
            else:
                return jsonify({'error': 'Client not found.'}), 404
    except mysql.connector.Error as db_err:
        logging.error(f"대표 코드 조회 실패: {db_err}")
        return jsonify({'error': 'Database error occurred.'}), 500
    finally:
        db.close()

def clean_virtual_account_number(van_number):
    """
    가상계좌번호에서 하이픈 제거
    """
    if isinstance(van_number, str):
        return van_number.replace('-', '').strip()
    return van_number

@app.route('/upload_bank_payments', methods=['GET', 'POST'])
def upload_bank_payments():
    form = UploadForm()
    if form.validate_on_submit():
        file = form.file.data
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # 업로드 디렉토리 생성
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            logging.debug(f"업로드 디렉토리 생성 또는 이미 존재: {app.config['UPLOAD_FOLDER']}")
            
            try:
                file.save(file_path)
                logging.info(f"업로드된 파일이 저장되었습니다: {file_path}")
            except Exception as e:
                logging.error(f"파일 저장 실패: {e}")
                flash('업로드된 파일을 저장하는 중 오류가 발생했습니다.', 'danger')
                return redirect(request.url)

            db = get_db_connection()
            if not db:
                flash('Database connection failed.', 'danger')
                return redirect(request.url)

            try:
                with db.cursor(dictionary=True) as cursor:
                    # 엑셀 파일을 헤더 없이 읽기
                    df = pd.read_excel(file_path, header=None)
                    logging.info(f"엑셀 파일을 성공적으로 읽었습니다: {filename}")
                    
                    # 헤더 행 찾기: 두 번째 열이 "입금일자"인 행 (첫 번째 열은 "No.")
                    header_row_index = None
                    for i, row in df.iterrows():
                        second_cell = str(row.iloc[1]).strip()
                        if second_cell == '입금일자':
                            header_row_index = i
                            break

                    if header_row_index is None:
                        logging.warning("엑셀 파일에서 헤더를 찾을 수 없습니다.")
                        flash('엑셀 파일에서 헤더를 찾을 수 없습니다.', 'danger')
                        return redirect(request.url)
                    
                    # 헤더가 있는 행부터 데이터 읽기
                    df = pd.read_excel(file_path, header=header_row_index)
                    logging.debug(f"헤더가 발견된 행: {header_row_index}")
                    logging.debug(f"데이터프레임 샘플:\n{df.head()}")

                    # 필요한 열만 선택 (불필요한 열은 무시)
                    required_columns = ['입금일자', '입금시간', '가상계좌번호', '입금금액']
                    if not all(col in df.columns for col in required_columns):
                        missing_cols = [col for col in required_columns if col not in df.columns]
                        logging.warning(f"엑셀 파일에 누락된 필드: {missing_cols}")
                        flash(f'엑셀 파일에 누락된 필드가 있습니다: {missing_cols}', 'danger')
                        return redirect(request.url)
                    
                    df = df[required_columns]

                    # 데이터 타입 강제 변환
                    df['입금일자'] = pd.to_datetime(df['입금일자'], errors='coerce').dt.date
                    df['입금시간'] = pd.to_datetime(df['입금시간'], format='%H:%M:%S', errors='coerce').dt.time
                    df['가상계좌번호'] = df['가상계좌번호'].apply(clean_virtual_account_number)
                    df['입금금액'] = pd.to_numeric(df['입금금액'].astype(str).str.replace(',', '', regex=True), errors='coerce')
                    
                    logging.debug(f"형변환 후 데이터프레임 샘플:\n{df.head()}")

                    # 데이터 삽입 준비
                    inserted_records = 0
                    for index, row in df.iterrows():
                        # 데이터 유효성 검사: '입금일자'와 '가상계좌번호'가 유효한지 확인
                        if pd.isna(row['입금일자']) or pd.isna(row['가상계좌번호']):
                            logging.debug(f"Row {index}은 '입금일자' 또는 '가상계좌번호'가 NaN이므로 무시됩니다.")
                            continue

                        payment_date = row['입금일자']
                        payment_time = row['입금시간']
                        virtual_account_number = row['가상계좌번호']
                        payment_amount = row['입금금액']

                        # 입금금액이 NaN인지 확인하고 기본값 설정
                        if pd.isna(payment_amount):
                            logging.warning(f"Row {index}의 '입금금액'이 NaN입니다. 기본값 0.00으로 설정합니다.")
                            payment_amount = Decimal('0.00')
                        else:
                            payment_amount = Decimal(payment_amount)

                        logging.debug(f"데이터 삽입 준비 - Row {index}: virtual_account_number={virtual_account_number}, payment_date={payment_date}, payment_amount={payment_amount}")

                        # ARBankAccountMaster 테이블에서 client_code, client_name, manager, collector_key, representative_code 조회
                        try:
                            cursor.execute("""
                                SELECT client_code, client_name, manager, collector_key, representative_code 
                                FROM ARBankAccountMaster 
                                WHERE REPLACE(hana_bank_virtual_account, '-', '') = %s 
                                LIMIT 1
                            """, (virtual_account_number,))
                            account_data = cursor.fetchone()
                            if account_data:
                                client_code = account_data['client_code']
                                client_name = account_data['client_name']
                                manager = account_data['manager']
                                collector_key = account_data['collector_key']
                                representative_code = account_data['representative_code']
                            else:
                                logging.warning(f"가상계좌번호 '{virtual_account_number}'에 해당하는 계좌를 ARBankAccountMaster에서 찾을 수 없습니다.")
                                flash(f"Row {index + 1}: 가상계좌번호 '{virtual_account_number}'에 해당하는 계좌를 찾을 수 없습니다.", 'warning')
                                continue
                        except mysql.connector.Error as db_err:
                            logging.error(f"ARBankAccountMaster에서 데이터 조회 실패 - Row {index}: {db_err}")
                            flash(f"Row {index + 1}: ARBankAccountMaster 조회 실패.", 'danger')
                            continue

                        
                        # ARBankPaymentDetails 테이블에 삽입하는 부분(반복문 내부)
                        # 중복 체크 쿼리 추가 start 0205
                        duplicate_query = """
                            SELECT COUNT(*) AS cnt
                            FROM ARBankPaymentDetails
                            WHERE payment_date = %s
                            AND payment_time = %s
                            AND client_code = %s
                            AND collector_key = %s
                            AND virtual_account_number = %s
                            AND payment_amount = %s
                        """
                        try:
                            cursor.execute(duplicate_query, (
                                payment_date,
                                payment_time,
                                client_code,
                                collector_key,
                                virtual_account_number,
                                payment_amount
                            ))
                            dup_result = cursor.fetchone()
                        except mysql.connector.Error as db_err:
                            logging.error(f"중복 체크 실패 - Row {index}: {db_err}")
                            flash(f"Row {index + 1}: 중복 체크 실패.", 'danger')
                            continue

                        if dup_result and dup_result['cnt'] > 0:
                            logging.info(f"Row {index}: 중복 데이터가 존재하여 삽입을 건너뜁니다.")
                            continue  # 중복 데이터면 다음 row로 넘어감
                        # 중복 체크 쿼리 추가 end
                        
                        
                        # ARBankPaymentDetails 테이블에 삽입
                        insert_bank_payment_query = """
                            INSERT INTO ARBankPaymentDetails (
                                payment_date, payment_time, client_code, collector_key, virtual_account_number, payment_amount
                            )
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """
                        try:
                            cursor.execute(insert_bank_payment_query, (
                                payment_date,
                                payment_time,
                                client_code,
                                collector_key,
                                virtual_account_number,
                                payment_amount
                            ))
                            logging.debug(f"ARBankPaymentDetails 삽입 성공 - Row {index}")
                        except mysql.connector.Error as db_err:
                            logging.error(f"ARBankPaymentDetails 삽입 실패 - Row {index}: {db_err}")
                            flash(f"Row {index + 1}: ARBankPaymentDetails 삽입 실패.", 'danger')
                            continue

                        # ARTransactionsLedger 테이블에 삽입
                        insert_ledger_query = """
                            INSERT INTO ARTransactionsLedger (
                                transaction_date, representative_code, client, outlet_name, debit, credit, cash_deposit
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """
                        try:
                            cursor.execute(insert_ledger_query, (
                                payment_date,
                                representative_code if representative_code else '',  # representative_code이 없을 경우 빈 값
                                client_code,
                                client_name,     # outlet_name에 client_name 사용 (필요 시 수정)
                                0,  # debit (식자재 매출)
                                payment_amount,       # credit
                                payment_amount   # cash_deposit
                            ))
                            logging.debug(f"ARTransactionsLedger 삽입 성공 - Row {index}")
                        except mysql.connector.Error as db_err:
                            logging.error(f"ARTransactionsLedger 삽입 실패 - Row {index}: {db_err}")
                            flash(f"Row {index + 1}: ARTransactionsLedger 삽입 실패.", 'danger')
                            continue

                        inserted_records += 1
                    db.commit()
                    logging.info(f"{inserted_records}개의 은행 입금 내역이 성공적으로 업로드되었습니다.")
                    flash(f'{inserted_records}개의 은행 입금 내역이 성공적으로 업로드되었습니다.', 'success')
                    return redirect(url_for('index'))
            except Exception as e:
                logging.error(f"업로드된 파일 처리 중 오류 발생: {e}")
                db.rollback()
                flash('파일을 처리하는 중 오류가 발생했습니다.', 'danger')
                return redirect(request.url)
            finally:
                db.close()

    return render_template('upload_bank_payments.html', form=form)

# 발주 내역 다량 업로드
@app.route('/upload_orders', methods=['GET', 'POST'])
def upload_orders():
    form = BulkUploadOrdersForm()
    if form.validate_on_submit():
        file = form.file.data
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # 업로드 디렉토리 생성 (존재하지 않을 경우)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            logging.debug(f"업로드 디렉토리 생성 또는 이미 존재: {app.config['UPLOAD_FOLDER']}")
            
            try:
                file.save(file_path)
                logging.info(f"업로드된 파일이 저장되었습니다: {file_path}")
            except Exception as e:
                logging.error(f"파일 저장 실패: {e}")
                flash('업로드된 파일을 저장하는 중 오류가 발생했습니다.', 'danger')
                return redirect(request.url)

            db = get_db_connection()
            if not db:
                flash('Database connection failed.', 'danger')
                return redirect(request.url)

            try:
                with db.cursor(dictionary=True) as cursor:
                    # 엑셀 파일 읽기 (헤더 없는 경우)
                    df = pd.read_excel(file_path, header=None)
                    logging.info(f"엑셀 파일을 성공적으로 읽었습니다: {filename}")
                    
                    # 헤더 행 찾기
                    header_row_index = None
                    for i, row in df.iterrows():
                        if all(col in row.values for col in ['order_date', 'client_code', 'order_amount', 'collector_key']):
                            header_row_index = i
                            break

                    if header_row_index is None:
                        logging.warning("엑셀 파일에서 헤더를 찾을 수 없습니다.")
                        flash('엑셀 파일에서 헤더를 찾을 수 없습니다.', 'danger')
                        return redirect(request.url)
                    
                    # 헤더가 있는 행부터 데이터 추출
                    df = pd.read_excel(file_path, header=header_row_index)
                    logging.debug(f"헤더가 발견된 행: {header_row_index}")
                    logging.debug(f"데이터프레임 샘플:\n{df.head()}")

                    # 필요한 컬럼 선택
                    required_columns = ['order_date', 'client_code', 'order_amount', 'collector_key']
                    if not all(col in df.columns for col in required_columns):
                        missing_cols = [col for col in required_columns if col not in df.columns]
                        logging.warning(f"엑셀 파일에 누락된 필드: {missing_cols}")
                        flash(f'엑셀 파일에 누락된 필드가 있습니다: {missing_cols}', 'danger')
                        return redirect(request.url)

                    df = df[required_columns]

                    # 데이터 타입 변환
                    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce').dt.date
                    df['client_code'] = df['client_code'].astype(str).str.strip()
                    df['order_amount'] = pd.to_numeric(df['order_amount'], errors='coerce').fillna(0)
                    df['collector_key'] = df['collector_key'].astype(str).str.strip()

                    logging.debug(f"형변환 후 데이터프레임 샘플:\n{df.head()}")

                    # 데이터 정제: 필수 컬럼 결측값 제거
                    df = df.dropna(subset=['order_date', 'client_code', 'collector_key'])

                    # ARClientMaster에서 필요한 데이터 가져오기
                    client_codes = df['client_code'].unique().tolist()
                    if not client_codes:
                        logging.warning("발주 내역에 client_code가 없습니다.")
                        flash('발주 내역에 client_code가 없습니다.', 'danger')
                        return redirect(request.url)

                    format_strings = ','.join(['%s'] * len(client_codes))
                    cursor.execute(f"SELECT client_code, representative_code, client_name, manager FROM ARClientMaster WHERE client_code IN ({format_strings})", tuple(client_codes))
                    clients = cursor.fetchall()
                    client_dict = {client['client_code']: client for client in clients}

                    # AROrderDetails 및 ARTransactionsLedger 데이터 준비
                    order_details_data = []
                    ledger_data = []
                    for index, row in df.iterrows():
                        client_code = row['client_code']
                        order_date = row['order_date']
                        order_amount = row['order_amount']
                        collector_key = row['collector_key']
                        
                        client = client_dict.get(client_code)
                        if not client:
                            logging.warning(f"client_code '{client_code}'가 ARClientMaster 테이블에 존재하지 않습니다.")
                            flash(f"Row {index + 1}: client_code '{client_code}'가 ARClientMaster 테이블에 존재하지 않습니다.", 'warning')
                            continue
                        representative_code = client['representative_code']
                        client_name = client['client_name']
                        manager = client['manager']

                        # AROrderDetails용 데이터 준비
                        order_details_data.append((
                            representative_code,
                            client_code,    # client_code 사용
                            client_name,
                            collector_key,
                            manager,
                            order_date,
                            order_amount
                        ))

                        # ARTransactionsLedger용 데이터 준비
                        ledger_data.append((
                            order_date,                # transaction_date
                            representative_code,       # representative_code
                            client_code,               # client
                            client_name,               # outlet_name
                            order_amount,              # debit
                            0,                          # credit
                            order_amount,              # food_material_sales
                            0,                         # royalty_sales
                            0,                         # advertising_fees
                            0,                         # other_sales
                            0,                         # cash_deposit
                            0,                         # meal_voucher_deposit
                            0,                         # delivery_fee
                            0,                         # card_deposit
                            0,                         # pos_usage_fee
                            0                          # receivables
                        ))

                    # AROrderDetails 테이블에 데이터 삽입
                    if order_details_data:
                        insert_order_query = """
                            INSERT INTO AROrderDetails (
                                representative_code, client_code, client_name, collector_key, manager, order_date, order_amount
                            )
                            VALUES (
                                %s, %s, %s, %s, %s, %s, %s
                            )
                        """
                        cursor.executemany(insert_order_query, order_details_data)
                        logging.debug(f"AROrderDetails 삽입 성공: {len(order_details_data)}건")
                    
                    # ARTransactionsLedger 테이블에 데이터 삽입
                    if ledger_data:
                        insert_ledger_query = """
                            INSERT INTO ARTransactionsLedger (
                                transaction_date,
                                representative_code,
                                client,
                                outlet_name,
                                debit,
                                credit,
                                food_material_sales,
                                royalty_sales,
                                advertising_fees,
                                other_sales,
                                cash_deposit,
                                meal_voucher_deposit,
                                delivery_fee,
                                card_deposit,
                                pos_usage_fee,
                                receivables
                            )
                            VALUES (
                                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                            )
                        """
                        cursor.executemany(insert_ledger_query, ledger_data)
                        logging.debug(f"ARTransactionsLedger 삽입 성공: {len(ledger_data)}건")

                    # 트랜잭션 커밋
                    db.commit()
                    inserted_records = len(order_details_data)
                    logging.info(f"{inserted_records}개의 발주 내역이 성공적으로 업로드되었습니다.")
                    flash(f'{inserted_records}개의 발주 내역이 성공적으로 업로드되었습니다.', 'success')
                    return redirect(url_for('index'))
            except Exception as e:
                logging.error(f"업로드된 파일 처리 중 오류 발생: {e}")
                db.rollback()
                flash('파일을 처리하는 중 오류가 발생했습니다.', 'danger')
                return redirect(request.url)
            finally:
                db.close()
    return render_template('upload_orders.html', form=form)


# 미수금액 조회
@app.route('/view_receivables', methods=['GET'])
def view_receivables():
    db = get_db_connection()
    if not db:
        flash('Database connection failed.', 'danger')
        return redirect(url_for('index'))

    try:
        with db.cursor(dictionary=True) as cursor:
            # 검색 파라미터 가져오기
            search_outlet = request.args.get('search_outlet', '').strip()

            # 수정된 쿼리: LEFT JOIN 및 집계 함수 추가
            query = """
                SELECT 
                    TRIM(UPPER(t.client)) AS client,
                    TRIM(UPPER(t.outlet_name)) AS outlet_name,
                    SUM(t.debit) AS total_debit,
                    SUM(t.credit) AS total_credit,
                    SUM(t.food_material_sales) AS total_food_material_sales,
                    SUM(t.royalty_sales) AS total_royalty_sales,
                    SUM(t.pos_usage_fee) AS total_pos_usage_fee,
                    SUM(t.cash_deposit) AS total_cash_deposit,
                    SUM(t.card_deposit) AS total_card_deposit,
                    SUM(t.debit) - SUM(t.credit) AS receivables,
                    IFNULL(MAX(m.deposit), 0) AS deposit
                FROM 
                    ARTransactionsLedger AS t
                LEFT JOIN 
                    ARClientMaster AS m ON TRIM(UPPER(t.client)) = TRIM(UPPER(m.client_code))
            """

            params = []

            # 검색 조건 추가
            if search_outlet:
                query += " WHERE TRIM(UPPER(t.outlet_name)) LIKE %s"
                params.append(f"%{search_outlet.upper().strip()}%")

            # GROUP BY 및 ORDER BY 추가
            query += """
                GROUP BY 
                    TRIM(UPPER(t.client)), 
                    TRIM(UPPER(t.outlet_name))
                ORDER BY 
                    client, outlet_name
            """

            cursor.execute(query, tuple(params))
            results = cursor.fetchall()

            # 합계 계산
            sum_total_debit = sum(row['total_debit'] for row in results) if results else 0
            sum_total_credit = sum(row['total_credit'] for row in results) if results else 0
            sum_food_material_sales = sum(row['total_food_material_sales'] for row in results) if results else 0
            sum_royalty_sales = sum(row['total_royalty_sales'] for row in results) if results else 0
            sum_pos_usage_fee = sum(row['total_pos_usage_fee'] for row in results) if results else 0
            sum_cash_deposit = sum(row['total_cash_deposit'] for row in results) if results else 0
            sum_card_deposit = sum(row['total_card_deposit'] for row in results) if results else 0
            sum_receivables = sum(row['receivables'] for row in results) if results else 0
            sum_deposit = sum(row['deposit'] for row in results) if results else 0  # 보증금 합계

            # 쿼리 결과 로그 출력
            logging.debug(f"미수금액 조회 결과: {results}")
            logging.info("미수금액 조회 성공")

            return render_template(
                'view_receivables.html', 
                results=results, 
                search_outlet=search_outlet,
                sum_total_debit=sum_total_debit,
                sum_total_credit=sum_total_credit,
                sum_food_material_sales=sum_food_material_sales,
                sum_royalty_sales=sum_royalty_sales,
                sum_pos_usage_fee=sum_pos_usage_fee,
                sum_cash_deposit=sum_cash_deposit,
                sum_card_deposit=sum_card_deposit,
                sum_receivables=sum_receivables,
                sum_deposit=sum_deposit  # 보증금 합계 전달
            )
    except mysql.connector.Error as db_err:
        logging.error(f"미수금액 조회 실패: {db_err}")
        flash('미수금액을 조회하는 중 오류가 발생했습니다.', 'danger')
        return redirect(url_for('index'))
    except Exception as e:
        logging.error(f"미수금액 조회 실패 (기타 오류): {e}")
        flash('미수금액을 조회하는 중 예상치 못한 오류가 발생했습니다.', 'danger')
        return redirect(url_for('index'))
    finally:
        db.close()


# view_daily_transactions 조회
def clean_decimal(value):
    """
    문자열에서 숫자와 소수점, 음수 기호만 남기고 제거한 후 Decimal로 변환합니다.
    """
    if isinstance(value, str):
        value = re.sub(r'[^\d.-]', '', value)  # 숫자, 소수점, 음수 기호만 남김
    try:
        return Decimal(value) if value else Decimal('0.00')
    except:
        return Decimal('0.00')

@app.route('/view_daily_transactions', methods=['GET'])
def view_daily_transactions():
    db = get_db_connection()
    if not db:
        flash('데이터베이스 연결에 실패했습니다.', 'danger')
        return redirect(url_for('index'))  # 'index' 라우트는 홈 페이지로 가정

    try:
        with db.cursor(dictionary=True) as cursor:
            # 검색 파라미터 가져오기
            search_outlet = request.args.get('search_outlet', '').strip()
            selected_year = request.args.get('year', datetime.now().year, type=int)
            selected_month = request.args.get('month', datetime.now().month, type=int)

            # 첫날과 마지막 날 계산
            try:
                first_day = datetime(selected_year, selected_month, 1).date()
                if selected_month == 12:
                    last_day = datetime(selected_year + 1, 1, 1).date()
                else:
                    last_day = datetime(selected_year, selected_month + 1, 1).date()
            except ValueError as ve:
                logging.error(f"날짜 계산 오류: {ve}")
                flash('유효하지 않은 날짜입니다.', 'danger')
                return redirect(url_for('index'))

            # SQL 쿼리 작성
            query = """
                SELECT 
                    A.client,
                    A.outlet_name,
                    C.collector_key,
                    C.manager,
            """

            # 동적 SQL 쿼리 생성: 각 일자별로 debit과 credit 합계를 계산
            day_sums = []
            for day in range(1, 32):
                day_sums.append(f"SUM(CASE WHEN DAY(A.transaction_date) = {day} THEN A.debit ELSE 0 END) AS `day_{day}_debit`")
                day_sums.append(f"SUM(CASE WHEN DAY(A.transaction_date) = {day} THEN A.credit ELSE 0 END) AS `day_{day}_credit`")

            # 총합계 계산 추가
            day_sums.append("SUM(A.debit) AS `total_debit`")
            day_sums.append("SUM(A.credit) AS `total_credit`")
            day_sums.append("(SUM(A.debit) - SUM(A.credit)) AS `total_receivables`")  # total_receivables 추가

            # SELECT 절에 모든 필드를 추가
            query += ",\n".join(day_sums)

            # FROM 절과 JOIN 절
            query += """
                FROM 
                    ARTransactionsLedger A
                JOIN 
                    ARClientMaster C ON A.client = C.client_code
            """

            # 검색 조건 추가
            conditions = []
            params = []
            if search_outlet:
                conditions.append("A.outlet_name LIKE %s")
                params.append(f"%{search_outlet}%")
            conditions.append("A.transaction_date >= %s AND A.transaction_date < %s")
            params.extend([first_day, last_day])

            if conditions:
                query += " WHERE " + " AND ".join(conditions)

            # GROUP BY 절 및 ORDER BY 절
            query += """
                GROUP BY 
                    A.client, A.outlet_name, C.collector_key, C.manager
                ORDER BY 
                    total_receivables DESC
            """

            logging.debug(f"실행할 쿼리: {query}")
            logging.debug(f"쿼리 파라미터: {params}")

            cursor.execute(query, tuple(params))
            results = cursor.fetchall()

            logging.debug(f"쿼리 결과: {results}")

            # 전달할 데이터 구조를 준비
            data = []
            sum_total_debit = Decimal('0.00')
            sum_total_credit = Decimal('0.00')
            sum_total_receivables = Decimal('0.00')  # 최종 합계에서 계산

            for row in results:
                # total_debit, total_credit, total_receivables을 Decimal으로 변환
                try:
                    total_debit = clean_decimal(row['total_debit'])
                    total_credit = clean_decimal(row['total_credit'])
                    total_receivables = clean_decimal(row['total_receivables'])
                except (ValueError, TypeError) as ve:
                    logging.error(f"데이터 변환 오류: {ve}")
                    total_debit = Decimal('0.00')
                    total_credit = Decimal('0.00')
                    total_receivables = Decimal('0.00')

                logging.debug(f"Row total_debit: {total_debit}, Row total_credit: {total_credit}, Row total_receivables: {total_receivables}")

                # 일별 데이터 수집
                day_data = {}
                for day in range(1, 32):
                    debit_key = f'day_{day}_debit'
                    credit_key = f'day_{day}_credit'
                    try:
                        debit_value = clean_decimal(row.get(debit_key, 0.0))
                        credit_value = clean_decimal(row.get(credit_key, 0.0))
                    except (ValueError, TypeError) as ve:
                        logging.error(f"일별 데이터 변환 오류: {ve}")
                        debit_value = Decimal('0.00')
                        credit_value = Decimal('0.00')

                    logging.debug(f"day_{day}_debit: {debit_value}, day_{day}_credit: {credit_value}")

                    # 데이터가 0보다 크면 포맷팅된 값, 아니면 '-'
                    day_data[f'day_{day}_debit'] = f"{debit_value:,.2f}" if debit_value > 0 else '-'
                    day_data[f'day_{day}_credit'] = f"{credit_value:,.2f}" if credit_value > 0 else '-'

                # 데이터 행 구성
                data_row = {
                    'client': row['client'],
                    'outlet_name': row['outlet_name'],
                    'collector_key': row['collector_key'],
                    'manager': row['manager'],
                    'total_debit': "{0:.2f}".format(total_debit) if total_debit > 0 else '0.00',
                    'total_credit': "{0:.2f}".format(total_credit) if total_credit > 0 else '0.00',
                    'total_receivables': "{0:.2f}".format(total_receivables) if total_receivables > 0 else '0.00',
                    'day_data': day_data
                }

                logging.debug(f"Data Row: {data_row}")

                # 누적 합계
                sum_total_debit += total_debit
                sum_total_credit += total_credit
                sum_total_receivables += total_receivables

                logging.debug(f"After adding row: sum_total_debit = {sum_total_debit}, sum_total_credit = {sum_total_credit}, sum_total_receivables = {sum_total_receivables}")

                data.append(data_row)

            # 최종 합계는 이미 누적됨
            logging.debug(f"Final Sum - Debit: {sum_total_debit}, Credit: {sum_total_credit}, Receivables: {sum_total_receivables}")

            # 전체 합계 포맷팅
            formatted_sum_total_debit = f"{sum_total_debit:,.2f}" if sum_total_debit > 0 else '-'
            formatted_sum_total_credit = f"{sum_total_credit:,.2f}" if sum_total_credit > 0 else '-'
            formatted_sum_total_receivables = f"{sum_total_receivables:,.2f}" if sum_total_receivables > 0 else '-'  # 전체 합계 포맷팅 추가

            return render_template(
                'view_daily_transactions.html',
                data=data,
                selected_year=selected_year,
                selected_month=selected_month,
                search_outlet=search_outlet,
                sum_total_debit=formatted_sum_total_debit,
                sum_total_credit=formatted_sum_total_credit,
                sum_total_receivables=formatted_sum_total_receivables  # 전체 합계 전달
            )
    except mysql.connector.Error as db_err:
        logging.error(f"데이터 조회 실패: {db_err}")
        flash('데이터를 조회하는 중 오류가 발생했습니다.', 'danger')
        return redirect(url_for('index'))
    except Exception as e:
        logging.error(f"예기치 않은 오류: {e}")
        flash('데이터를 처리하는 중 오류가 발생했습니다.', 'danger')
        return redirect(url_for('index'))
    finally:
        db.close()

# ------------------------
# 새로운 라우트: 웹발주 엑셀다운로드
# ------------------------
@app.route('/download_web_order_excel', methods=['GET'])
def download_web_order_excel():
    try:
        excel_file_path = etl_process()
        
        if excel_file_path is None:
            logging.warning("엑셀 파일이 생성되지 않았습니다. 데이터가 없거나 처리 과정에서 문제가 발생했습니다.")
            flash("엑셀 파일을 생성할 데이터가 없습니다.", "warning")
            return redirect(url_for('index'))
        
        # 엑셀 파일이 저장된 디렉토리와 파일명을 분리
        directory, filename = os.path.split(excel_file_path)
        
        if not os.path.exists(excel_file_path):
            logging.error(f"엑셀 파일이 존재하지 않습니다: {excel_file_path}")
            flash("엑셀 파일이 존재하지 않습니다.", "danger")
            return redirect(url_for('index'))
        
        # 다운로드 링크 생성
        return send_from_directory(directory, filename, as_attachment=True, download_name=filename)
    except Exception as e:
        logging.error(f"웹발주 엑셀 다운로드 중 오류 발생: {e}")
        flash(f"웹발주 엑셀 다운로드 중 오류가 발생했습니다: {e}", 'danger')
        return redirect(url_for('index'))

# 기존의 다른 라우트들이 이미 정의되어 있으므로 추가로 수정할 필요 없음

# 이미 다른 라우트들에서 정의된 download_file 라우트는 필요 없을 수도 있지만, 
# 만약 필요한 경우 그대로 유지
@app.route('/download/<filename>')
def download_file(filename):
    directory = os.path.join(os.getcwd(), 'excel_output')
    return send_from_directory(directory, filename, as_attachment=True, download_name=filename)


# ------------------------
# 8. 추가 라우트: 거래명세표 엑셀 및 PDF 다운로드 폼 및 처리
# ------------------------

@app.route('/download_orders_excel_form', methods=['GET'])
def download_orders_excel_form():
    """
    거래명세표 엑셀 및 PDF 다운로드 폼을 표시하는 라우트.
    """
    form = DownloadOrdersForm()
    return render_template('download_orders_form.html', form=form)

@app.route('/download_orders_excel', methods=['POST'])
def download_orders_excel():
    """
    사용자로부터 받은 order_date에 따라 거래명세표 엑셀 및 PDF 파일을 생성하고 다운로드하는 라우트.
    """
    form = DownloadOrdersForm()
    if form.validate_on_submit():
        order_date = form.order_date.data.strftime('%Y-%m-%d')
        try:
            # ETL 프로세스 실행하여 엑셀 및 PDF 파일 생성
            file_paths = export_orders_to_files(order_date)
            
            if not file_paths:
                flash("파일이 생성되지 않았습니다.", "danger")
                return redirect(url_for('download_orders_excel_form'))
            
            # 생성된 모든 파일을 ZIP으로 압축하여 다운로드
            with tempfile.TemporaryDirectory() as tmpdirname:
                zip_path = os.path.join(tmpdirname, f"거래명세표_{order_date}.zip")
                with ZipFile(zip_path, 'w') as zipf:
                    for file_path in file_paths:
                        zipf.write(file_path, os.path.basename(file_path))
                return send_from_directory(
                    directory=tmpdirname,
                    path=os.path.basename(zip_path),
                    as_attachment=True,
                    download_name=f"거래명세표_{order_date}.zip"
                )
            
        except Exception as e:
            logging.error(f"거래명세표 다운로드 중 오류 발생: {e}")
            flash(f"거래명세표 다운로드 중 오류가 발생했습니다: {e}", 'danger')
            return redirect(url_for('download_orders_excel_form'))
    else:
        # 폼 검증 실패 시
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text} - {error}", 'danger')
        return redirect(url_for('download_orders_excel_form'))

# ------------------------
# 9. 공급자 정보 (중복 제거)
# ------------------------

# 공급자 정보는 이미 SUPPLIER_INFO 상수로 정의되어 있으므로 추가 정의 제거

# ------------------------
# 10. 기타 헬퍼 함수 및 로깅 설정
# ------------------------


def load_excel_template():
    """
    엑셀 템플릿을 로드합니다.
    """
    if not os.path.exists(TEMPLATE_FILE):
        logging.error(f"엑셀 템플릿 파일이 존재하지 않습니다: {TEMPLATE_FILE}")
        raise FileNotFoundError(f"엑셀 템플릿 파일이 존재하지 않습니다: {TEMPLATE_FILE}")
    
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    logging.info(f"엑셀 템플릿을 성공적으로 로드했습니다: {TEMPLATE_FILE}")
    return wb, ws

def insert_cell_value(ws, row, column, value):
    """
    지정된 셀에 값을 삽입합니다. 병합된 셀일 경우 첫 번째 셀에만 값을 할당합니다.
    
    :param ws: Worksheet 객체
    :param row: 행 번호
    :param column: 열 번호
    :param value: 할당할 값
    """
    cell = ws.cell(row=row, column=column)
    
    if cell.coordinate in ws.merged_cells:
        # 병합된 셀인 경우, 병합된 영역의 첫 번째 셀에 값 할당
        for merged_cell in ws.merged_cells.ranges:
            if cell.coordinate in merged_cell:
                main_cell = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col)
                main_cell.value = value
                break
    else:
        # 병합되지 않은 셀인 경우, 직접 값 할당
        cell.value = value

# 헬퍼 함수: 파일 확장자 확인 (기존)
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def insert_task(task_id, status):
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            query = "INSERT INTO background_tasks (task_id, status, created_at, updated_at) VALUES (%s, %s, NOW(), NOW())"
            cursor.execute(query, (task_id, status))
            db.commit()
    except Exception as e:
        logging.error(f"Insert task error: {e}", exc_info=True)
    finally:
        db.close()

def update_task_status(task_id, status, result):
    db = get_db_connection()
    try:
        with db.cursor() as cursor:
            query = "UPDATE background_tasks SET status=%s, result=%s, updated_at=NOW() WHERE task_id=%s"
            cursor.execute(query, (status, result, task_id))
            db.commit()
    except Exception as e:
        logging.error(f"Update task error: {e}", exc_info=True)
    finally:
        db.close()

def get_task(task_id):
    db = get_db_connection()
    try:
        with db.cursor(dictionary=True) as cursor:
            query = "SELECT * FROM background_tasks WHERE task_id=%s"
            cursor.execute(query, (task_id,))
            task = cursor.fetchone()
            return task
    except Exception as e:
        logging.error(f"Get task error: {e}", exc_info=True)
        return None
    finally:
        db.close()



# ---------------------------
# 신규 폼 클래스: DownloadClientOrdersForm
# ---------------------------
class DownloadClientOrdersForm(FlaskForm):
    # 필수 검증기를 제거하여 빈 값도 허용 (전체 매출처 의미)
    client_code = SelectField('매출처', choices=[('', '전체매출처')])
    from_date = DateField('시작일', format='%Y-%m-%d', validators=[DataRequired()])
    to_date = DateField('종료일', format='%Y-%m-%d', validators=[DataRequired()])
    submit = SubmitField('거래명세표 생성')

# ------------------------
# 10. 애플리케이션 실행
# ------------------------

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
