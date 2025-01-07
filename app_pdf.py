import os
import re
import logging
import win32com.client
import pythoncom
import traceback
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def setup_export_logging():
    """
    엑셀 내보내기 전용 로깅 설정 함수.
    """
    logging.info("엑셀 내보내기 프로세스 시작.")

def get_sqlalchemy_engine():
    """
    SQLAlchemy 엔진을 생성하는 함수.
    기존 MySQL 연결 정보를 활용합니다.
    """
    try:
        engine = create_engine(
            f"mysql+pymysql://{os.getenv('DB_USER', 'nolboo')}:{os.getenv('DB_PASSWORD', '2024!puser')}@"
            f"{os.getenv('DB_HOST', '175.196.7.45')}/{os.getenv('DB_NAME', 'nolboo')}?charset=utf8mb4"
        )
        logging.info("SQLAlchemy 엔진이 성공적으로 생성되었습니다.")
        return engine
    except Exception as e:
        logging.error(f"SQLAlchemy 엔진 생성 오류: {e}")
        return None

def preprocess_data(df):
    """
    MySQL에서 조회한 데이터 전처리 함수.
    필요한 컬럼만 선택하고 데이터 타입 변환.
    """
    # 필요한 컬럼 확인
    required_columns = list(COLUMN_MAPPING.values()) + ['full_name', 'reg_no', 'president', 'address1']  # 추가 컬럼
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.error(f"MySQL 데이터에 누락된 컬럼이 있습니다: {missing_columns}")
        raise ValueError(f"MySQL 데이터에 누락된 컬럼이 있습니다: {missing_columns}")
    
    # 필요한 컬럼만 선택
    df = df[required_columns]
    
    # 데이터 타입 변환
    # order_date: 문자열 (날짜 형식)
    df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce').dt.strftime('%Y-%m-%d')
    
    # 수량 관련 컬럼: Decimal(15,2)
    decimal_columns = ['qty', 'cal_qty', 'unit_price', 'order_amount', 'vat', 'total_amount']
    for col in decimal_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).round(2)
    
    # 기타 문자열 필드: strip() 적용
    string_fields = ['rep_code', 'rep_name', 'client_code', 'client_name', 
                        'item_code', 'item_name', 'cond', 'unit', 'tax']
    for field in string_fields:
        df[field] = df[field].astype(str).str.strip()
    
    # 공급받는자 정보 (매출처 정보)
    df['full_name'] = df['full_name'].fillna('-').astype(str).str.strip()
    df['reg_no'] = df['reg_no'].fillna('-').astype(str).str.strip()
    df['president'] = df['president'].fillna('-').astype(str).str.strip()
    df['address1'] = df['address1'].fillna('-').astype(str).str.strip()
    
    logging.info("데이터 전처리 및 형변환이 완료되었습니다.")
    
    return df

def load_excel_template():
    """
    엑셀 템플릿을 로드합니다.
    """
    if not os.path.exists(TEMPLATE_FILE):
        logging.error(f"엑셀 템플릿 파일이 존재하지 않습니다: {TEMPLATE_FILE}")
        raise FileNotFoundError(f"엑셀 템플릿 파일이 존재하지 않습니다: {TEMPLATE_FILE}")
    
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    logging.info(f"엑셀 템플릿을 성공적으로 로드했습니다: {TEMPLATE_FILE}")
    return wb, ws

def insert_cell_value(ws, row, column, value):
    """
    지정된 셀에 값을 삽입합니다. 병합된 셀일 경우 첫 번째 셀에만 값을 할당합니다.
    
    :param ws: Worksheet 객체
    :param row: 행 번호
    :param column: 열 번호
    :param value: 할당할 값
    """
    cell = ws.cell(row=row, column=column)
    
    if cell.coordinate in ws.merged_cells:
        # 병합된 셀인 경우, 병합된 영역의 첫 번째 셀에 값 할당
        for merged_cell in ws.merged_cells.ranges:
            if cell.coordinate in merged_cell:
                main_cell = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col)
                main_cell.value = value
                break
    else:
        # 병합되지 않은 셀인 경우, 직접 값 할당
        cell.value = value

def insert_data_to_excel(wb, ws, supplier_info, client_info, order_date, data_rows):
    """
    엑셀 템플릿에 데이터를 삽입하는 함수.
    """
    # 공급자 정보 삽입 (고정된 위치)
    ws['G3'] = supplier_info['등록번호']
    ws['G4'] = supplier_info['상호 (법인명)']
    ws['G5'] = supplier_info['성명']
    ws['G6'] = supplier_info['주소']
    
    # 공급받는자 정보 삽입 (동적 위치)
    # 사업자 등록번호를 10개의 셀에 한 글자씩 할당
    reg_no = client_info.get('reg_no', '-').replace('-', '')  # 하이픈 제거
    reg_no = reg_no.ljust(10, '-')[:10]  # 10자리 맞추기
    
    reg_no_cells = ['V3', 'W3', 'X3', 'Z3', 'AA3', 'AC3', 'AD3', 'AE3', 'AF3', 'AG3']
    
    for cell, char in zip(reg_no_cells, reg_no):
        # 셀 문자열에서 열 문자와 행 번호 분리
        column_letters = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell)))
        column_num = column_index_from_string(column_letters)
        insert_cell_value(ws, row_num, column_num, char)
    
    # 다른 공급받는자 정보 할당
    ws['V4'] = f"{client_info.get('full_name', '-')}"
    ws['V5'] = f"{client_info.get('president', '-')}"
    ws['V6'] = f"{client_info.get('address1', '-')}"
    
    # rep_name을 U2에 할당
    if data_rows:
        rep_name = data_rows[0].get('rep_name', '-')
        ws['U2'] = rep_name
    else:
        ws['U2'] = '-'
    
    # 거래일자 삽입
    ws['A2'] = f"{order_date}"
    
    # 데이터 삽입 시작 행 (예: 8행부터)
    current_row = 8
    
    # 합계 계산을 위한 변수 초기화
    total_order_amount = 0.0
    total_vat = 0.0
    
    for row in data_rows:
        try:
            # 품명 (컬럼 1)
            insert_cell_value(ws, current_row, 1, row['item_name'])
            # 단위 (컬럼 14)
            insert_cell_value(ws, current_row, 14, row['unit'])
            # 수량 (컬럼 17)
            insert_cell_value(ws, current_row, 17, float(row['qty']))
            # 단가 (컬럼 21)
            insert_cell_value(ws, current_row, 21, float(row['unit_price']))
            # 금액 (컬럼 26)
            insert_cell_value(ws, current_row, 26, float(row['order_amount']))
            # 세액 (컬럼 30)
            insert_cell_value(ws, current_row, 30, float(row['vat']))
            
            # 합계 누적
            total_order_amount += float(row['order_amount'])
            total_vat += float(row['vat'])
            
            current_row +=1
        except Exception as e:
            logging.error(f"데이터 삽입 중 오류 발생 (행: {current_row}): {e}")
            continue  # 다음 행으로 넘어가기
    
    # 공백 행 추가
    current_row +=1
    
    # 합계 값을 특정 셀에 할당
    try:
        ws['Z42'] = total_order_amount
        ws['AD42'] = total_vat
        ws['M43'] = total_order_amount + total_vat  # M43로 변경
        
        # 합계 셀에 숫자 형식 적용 (콤마 구분 및 소수점 2자리)
        ws['Z42'].number_format = '#,##0'
        ws['AD42'].number_format = '#,##0'
        ws['M43'].number_format = '#,##0'
        
        logging.info(f"합계 - order_amount: {total_order_amount}, vat: {total_vat}, total: {total_order_amount + total_vat}이(가) Z42, AD42, M43 셀에 할당되었습니다.")
    except Exception as e:
        logging.error(f"합계 셀 할당 중 오류 발생: {e}")
    
    logging.info(f"엑셀에 {len(data_rows)}개의 데이터가 삽입되었습니다.")

def generate_excel_file(wb, ws, client_name, order_date, group, output_folder):
    """
    단일 거래처와 일자에 대한 엑셀 파일을 생성하고 저장하며, PDF로도 변환합니다.
    """
    try:
        # 공급받는자 정보 가져오기
        client_info = {
            'full_name': group['full_name'].iloc[0],
            'reg_no': group['reg_no'].iloc[0],
            'president': group['president'].iloc[0],
            'address1': group['address1'].iloc[0]
        }
        
        # 데이터 삽입
        insert_data_to_excel(wb, ws, SUPPLIER_INFO, client_info, order_date, group.to_dict('records'))
        
        # 파일명 생성: 거래명세표_고객명_배송일자.xlsx
        sanitized_client_name = re.sub(r'[\\/*?:"<>|]', "_", client_info.get('full_name', 'Unknown'))
        sanitized_order_date = order_date.replace('-', '')
        excel_filename = f"거래명세표_{sanitized_client_name}_{sanitized_order_date}.xlsx"
        excel_path = os.path.join(output_folder, excel_filename)
        
        # 엑셀 파일 저장
        wb.save(excel_path)
        logging.info(f"엑셀 파일이 성공적으로 저장되었습니다: {excel_path}")
        
        # PDF 파일 경로 생성
        pdf_filename = excel_filename.replace(".xlsx", ".pdf")
        pdf_path = os.path.join(output_folder, pdf_filename)
        
        # Excel 파일을 PDF로 변환
        success = excel_to_pdf(excel_path, pdf_path)
        if success:
            logging.info(f"PDF 파일이 성공적으로 생성되었습니다: {pdf_path}")
        else:
            logging.error(f"PDF 파일 생성에 실패했습니다: {pdf_path}")
        
        return excel_path, pdf_path
    except Exception as e:
        logging.error(f"엑셀 및 PDF 파일 생성 중 오류 발생 (고객명: {client_name}, 배송일자: {order_date}): {e}")
        logging.error(traceback.format_exc())
        raise e

def fetch_data(engine, target_date):
    """
    SQLAlchemy 엔진을 사용하여 데이터베이스에서 특정 날짜의 데이터를 조회합니다.
    
    :param engine: SQLAlchemy 엔진 객체
    :param target_date: 처리할 배송일자 (YYYY-MM-DD)
    :return: Pandas DataFrame
    """
    if engine is None:
        logging.error("SQLAlchemy 엔진이 생성되지 않았습니다.")
        return None
    
    try:
        query = f"""
            SELECT 
                a.order_date,
                a.rep_code,
                a.rep_name,
                a.client_code,
                a.client_name,
                a.item_code,
                a.item_name,
                a.cond,
                a.unit,
                a.qty,
                a.cal_qty,
                a.unit_price,
                a.order_amount,
                a.vat,
                a.total_amount,
                a.tax,
                c.full_name,       -- 매출처명
                c.reg_no,          -- 매출처 등록번호
                c.president,       -- 매출처 성명
                c.address1         -- 매출처 주소
            FROM 
                {AR_ORDER_DETAILS_ITEM_TABLE} a
            LEFT JOIN 
                {CM_CHAIN_TABLE} c ON a.client_code COLLATE utf8mb4_unicode_ci = c.chain_no COLLATE utf8mb4_unicode_ci
            WHERE
                a.order_date = '{target_date}'
            ORDER BY 
                a.client_code, a.order_date
        """
        df = pd.read_sql(query, engine)
        logging.info(f"MySQL에서 데이터를 성공적으로 조회했습니다. 총 {len(df)}개의 레코드.")
        return df
    except Exception as e:
        logging.error(f"데이터 조회 중 오류 발생: {e}")
        return None

def export_orders_to_excel(order_date, output_folder='output_files'):
    """
    특정 배송일자에 대한 거래명세표 엑셀 및 PDF 파일을 생성하고 저장합니다.
    
    :param order_date: 처리할 배송일자 (YYYY-MM-DD)
    :param output_folder: 엑셀 및 PDF 파일을 저장할 폴더 경로
    :return: 생성된 엑셀 및 PDF 파일의 경로 리스트
    """
    setup_export_logging()
    
    # 출력 폴더가 존재하지 않으면 생성
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        logging.info(f"출력 폴더를 생성했습니다: {output_folder}")
    
    # SQLAlchemy 엔진 생성
    engine = get_sqlalchemy_engine()
    if engine is None:
        logging.error("SQLAlchemy 엔진 생성에 실패하여 ETL 프로세스를 종료합니다.")
        raise ConnectionError("SQLAlchemy 엔진 생성에 실패했습니다.")
    
    # 데이터베이스에서 데이터 조회
    df = fetch_data(engine, order_date)
    if df is None or df.empty:
        logging.error("데이터 조회에 실패하거나 데이터가 없습니다.")
        raise ValueError("데이터 조회에 실패하거나 데이터가 없습니다.")
    
    # 그룹화: client_code
    grouped = df.groupby(['client_code'])
    logging.info(f"배송일자 {order_date}에 대한 데이터는 {len(grouped)}개의 거래처로 나뉩니다.")
    
    excel_file_paths = []  # 생성된 엑셀 파일 경로를 저장할 리스트
    pdf_file_paths = []    # 생성된 PDF 파일 경로를 저장할 리스트
    
    for client_code, group in grouped:
        client_name = group['client_name'].iloc[0]
        logging.info(f"Processing client_code: {client_code}, client_name: {client_name}")
        
        # 데이터 전처리
        try:
            df_processed = preprocess_data(group)
        except Exception as e:
            logging.error(f"데이터 전처리 중 오류 발생 (client_code: {client_code}): {e}")
            continue  # 다음 그룹으로 넘어가기
        
        # 엑셀 템플릿 로드
        try:
            wb, ws = load_excel_template()
        except Exception as e:
            logging.error(f"엑셀 템플릿 로드 중 오류 발생: {e}")
            continue  # 다음 그룹으로 넘어가기
        
        # 엑셀 및 PDF 파일 생성
        try:
            excel_path, pdf_path = generate_excel_file(wb, ws, client_name, order_date, df_processed, output_folder)
            excel_file_paths.append(excel_path)
            if os.path.exists(pdf_path):
                pdf_file_paths.append(pdf_path)
        except Exception as e:
            logging.error(f"엑셀/PDF 파일 생성 중 오류 발생 (client_code: {client_code}, order_date: {order_date}): {e}")
            continue  # 다음 그룹으로 넘어가기
    
    logging.info("모든 엑셀 및 PDF 파일이 성공적으로 생성되었습니다.")
    return excel_file_paths + pdf_file_paths
